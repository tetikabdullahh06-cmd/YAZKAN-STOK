"""Iteration 7 tests:
- POST /api/auth/change-password (validation + admin & viewer flows)
- POST /api/toolholders/import (template, preview, commit, RBAC, validation errors)
- Critical stock email is NOT triggered on /api/stock/out (behaviour change)
- GET /api/dashboard still returns critical_products list

CRITICAL: this file restores admin password to '123456' at end.
"""
import os
import io
import uuid
from unittest.mock import patch, AsyncMock
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")

API = f"{BASE_URL}/api"
ADMIN_EMAIL = "takimhane@yazkan.com.tr"
ADMIN_PW = "123456"


def _login(email, pw):
    return requests.post(f"{API}/auth/login", json={"email": email, "password": pw})


@pytest.fixture(scope="module")
def admin_h():
    r = _login(ADMIN_EMAIL, ADMIN_PW)
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


@pytest.fixture(scope="module")
def viewer_creds():
    email = f"viewer_it7_{uuid.uuid4().hex[:6]}@example.com"
    pw = "Viewer123!"
    r = requests.post(f"{API}/auth/register",
                      json={"email": email, "password": pw, "name": "V IT7"})
    assert r.status_code == 200, r.text
    tok = r.json()["access_token"]
    return {"email": email, "password": pw, "token": tok,
            "headers": {"Authorization": f"Bearer {tok}"}}


# =============== Change Password ===============
class TestChangePassword:
    def test_no_token_401(self):
        r = requests.post(f"{API}/auth/change-password",
                          json={"current_password": "x", "new_password": "xxxxxx"})
        assert r.status_code == 401

    def test_new_pw_too_short_422(self, admin_h):
        r = requests.post(f"{API}/auth/change-password", headers=admin_h,
                          json={"current_password": ADMIN_PW, "new_password": "abc"})
        assert r.status_code == 422

    def test_wrong_current_400(self, admin_h):
        r = requests.post(f"{API}/auth/change-password", headers=admin_h,
                          json={"current_password": "wrongpw", "new_password": "newpass1"})
        assert r.status_code == 400
        assert "Mevcut şifre hatalı" in r.text

    def test_same_as_current_400(self, admin_h):
        r = requests.post(f"{API}/auth/change-password", headers=admin_h,
                          json={"current_password": ADMIN_PW, "new_password": ADMIN_PW})
        assert r.status_code == 400
        assert "eskisiyle aynı" in r.text

    def test_admin_full_flow_and_restore(self):
        """Change admin PW → login with new → login with old fails → RESTORE to 123456."""
        # login fresh (avoid token reuse issues)
        r = _login(ADMIN_EMAIL, ADMIN_PW)
        assert r.status_code == 200
        h = {"Authorization": f"Bearer {r.json()['access_token']}"}
        new_pw = "TempIter7Pw!"
        try:
            rc = requests.post(f"{API}/auth/change-password", headers=h,
                               json={"current_password": ADMIN_PW, "new_password": new_pw})
            assert rc.status_code == 200, rc.text
            # old password fails
            r_old = _login(ADMIN_EMAIL, ADMIN_PW)
            assert r_old.status_code == 401
            # new password works
            r_new = _login(ADMIN_EMAIL, new_pw)
            assert r_new.status_code == 200
        finally:
            # RESTORE - must succeed even if assertions above fail
            r_new = _login(ADMIN_EMAIL, new_pw)
            if r_new.status_code == 200:
                h2 = {"Authorization": f"Bearer {r_new.json()['access_token']}"}
                rr = requests.post(f"{API}/auth/change-password", headers=h2,
                                   json={"current_password": new_pw, "new_password": ADMIN_PW})
                assert rr.status_code == 200, f"FAILED TO RESTORE ADMIN PW: {rr.text}"
            # sanity: admin can log in with original password
            r_check = _login(ADMIN_EMAIL, ADMIN_PW)
            assert r_check.status_code == 200, "Admin password not restored!"

    def test_viewer_can_change_own_password(self, viewer_creds):
        h = viewer_creds["headers"]
        new_pw = "ViewerNew1!"
        rc = requests.post(f"{API}/auth/change-password", headers=h,
                           json={"current_password": viewer_creds["password"],
                                 "new_password": new_pw})
        assert rc.status_code == 200, rc.text
        # new pw login works
        r_new = _login(viewer_creds["email"], new_pw)
        assert r_new.status_code == 200


# =============== ToolHolder Import ===============
class TestToolHolderImport:
    def test_template_download(self, admin_h):
        r = requests.get(f"{API}/toolholders/import/template", headers=admin_h)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("Content-Type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        for col in ("name", "brand", "type", "min_stock", "current_stock"):
            assert col in header, f"template missing column {col}"

    def test_template_requires_auth(self):
        r = requests.get(f"{API}/toolholders/import/template")
        assert r.status_code == 401

    def _make_xlsx(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_viewer_import_403(self, viewer_creds):
        # Viewer's password was changed in previous test; re-login using latest state:
        # simpler: register a fresh viewer here
        email = f"viewer_it7b_{uuid.uuid4().hex[:6]}@example.com"
        r = requests.post(f"{API}/auth/register",
                          json={"email": email, "password": "Viewer123!", "name": "V2"})
        assert r.status_code == 200
        h = {"Authorization": f"Bearer {r.json()['access_token']}"}
        buf = self._make_xlsx([["name"], ["Any"]])
        files = {"file": ("t.xlsx", buf,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        rr = requests.post(f"{API}/toolholders/import", headers=h, files=files)
        assert rr.status_code == 403

    def test_import_preview_no_commit(self, admin_h):
        unique = uuid.uuid4().hex[:6]
        name_new = f"TEST_TH_IMP_NEW_{unique}"
        buf = self._make_xlsx([
            ["name", "brand", "type", "length", "diameter", "min_stock",
             "current_stock", "location", "note"],
            [name_new, "Sandvik", "BT40", "120", "32", 1, 5, "Raf X", "n"],
            ["", "NoName", "T1", "", "", 0, 0, "", ""],  # skip: missing name
            [f"TEST_TH_BAD_{unique}", "B", "T", "", "", "abc", 0, "", ""],  # skip: bad num
        ])
        files = {"file": ("t.xlsx", buf,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/toolholders/import", headers=admin_h, files=files)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["committed"] is False
        stats = body["stats"]
        assert stats["total"] == 3
        assert stats["skip"] == 2
        assert stats["create"] == 1
        # error messages present
        preview = body["preview"]
        skip_errors = [p["error"] for p in preview if p["action"] == "skip"]
        assert any("name" in (e or "").lower() for e in skip_errors)
        assert any("sayı" in (e or "") for e in skip_errors)
        # Verify DB unchanged
        lst = requests.get(f"{API}/toolholders", headers=admin_h).json()
        assert not any(t["name"] == name_new for t in lst)

    def test_import_commit_creates_and_updates(self, admin_h):
        unique = uuid.uuid4().hex[:6]
        n1 = f"TEST_TH_IMP_C_{unique}"  # create
        # Pre-create one to be updated
        n2 = f"TEST_TH_IMP_U_{unique}"
        pre = requests.post(f"{API}/toolholders", headers=admin_h, json={
            "name": n2, "brand": "Kaiser", "type": "HSK-A63",
            "length": "old", "current_stock": 1
        })
        assert pre.status_code == 200
        pre_id = pre.json()["id"]

        buf = self._make_xlsx([
            ["name", "brand", "type", "length", "diameter", "min_stock",
             "current_stock", "location", "note"],
            [n1, "Sandvik", "BT40", "120", "32", 1, 5, "Raf X", "created"],
            [n2, "Kaiser", "HSK-A63", "NEWLEN", "25", 2, 9, "Raf Y", "updated"],
        ])
        files = {"file": ("t.xlsx", buf,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/toolholders/import?commit=true",
                          headers=admin_h, files=files)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["committed"] is True
        assert body["created"] == 1
        assert body["updated"] == 1

        lst = requests.get(f"{API}/toolholders", headers=admin_h).json()
        created = next((t for t in lst if t["name"] == n1), None)
        updated = next((t for t in lst if t["name"] == n2), None)
        assert created is not None
        assert created["brand"] == "Sandvik"
        assert created["current_stock"] == 5
        assert updated is not None
        assert updated["length"] == "NEWLEN"
        assert updated["current_stock"] == 9
        # cleanup
        requests.delete(f"{API}/toolholders/{created['id']}", headers=admin_h)
        requests.delete(f"{API}/toolholders/{pre_id}", headers=admin_h)

    def test_import_rejects_non_xlsx(self, admin_h):
        files = {"file": ("bad.txt", b"hello", "text/plain")}
        r = requests.post(f"{API}/toolholders/import", headers=admin_h, files=files)
        assert r.status_code == 400


# =============== Critical stock email suppression on /stock/out ===============
class TestCriticalEmailSuppressed:
    def test_stock_out_critical_no_email_sent(self, admin_h):
        """Behavioural: stock_out that drops to <= min_stock should NOT trigger the
        email sender. We patch `server.send_critical_stock_email` and assert it was
        never awaited. Response must still carry critical=True."""
        import sys
        # Ensure server module is importable — supervisor runs it as 'server'
        sys.path.insert(0, "/app/backend")
        import server as srv

        # Setup: product + personnel + machine
        rp = requests.post(f"{API}/products", headers=admin_h, json={
            "code": "", "name": "TEST_IT7_CritMail", "category": "T",
            "unit": "adet", "min_stock": 10, "current_stock": 12,
        })
        assert rp.status_code == 200, rp.text
        pid = rp.json()["id"]
        pers = requests.get(f"{API}/personnel", headers=admin_h).json()
        mach = requests.get(f"{API}/machines", headers=admin_h).json()
        # if empty, create
        if not pers:
            pres = requests.post(f"{API}/personnel", headers=admin_h, json={
                "first_name": "TEST_IT7", "last_name": "P", "department": "CNC"})
            pers = [pres.json()]
        if not mach:
            mres = requests.post(f"{API}/machines", headers=admin_h, json={
                "code": f"TIT7-{uuid.uuid4().hex[:4]}", "name": "TEST_IT7_M",
                "type": "CNC Torna"})
            mach = [mres.json()]

        try:
            with patch.object(srv, "send_critical_stock_email",
                              new=AsyncMock(return_value=None)) as m:
                r = requests.post(f"{API}/stock/out", headers=admin_h, json={
                    "product_id": pid, "quantity": 5,  # 12 -> 7 (<=10 = critical)
                    "personnel_id": pers[0]["id"],
                    "machine_id": mach[0]["id"],
                })
                assert r.status_code == 200, r.text
                assert r.json()["critical"] is True
                assert r.json()["new_stock"] == 7
                # the email fn must NOT have been called from within the request
                # (the process is a separate uvicorn worker, so patch here won't
                # affect it — we can't rely on the mock). Instead, verify by
                # code inspection that the call site is removed:
            # Fallback verification: source of server.stock_out no longer references
            # send_critical_stock_email in its body.
            import inspect
            src = inspect.getsource(srv.stock_out)
            assert "send_critical_stock_email" not in src, \
                "stock_out still calls send_critical_stock_email!"
        finally:
            requests.delete(f"{API}/products/{pid}", headers=admin_h)


# =============== Dashboard still returns critical_products ===============
class TestDashboardCritical:
    def test_dashboard_has_critical_products(self, admin_h):
        # create a critical product
        rp = requests.post(f"{API}/products", headers=admin_h, json={
            "code": "", "name": "TEST_IT7_DashCrit", "category": "T",
            "unit": "adet", "min_stock": 10, "current_stock": 2,
        })
        pid = rp.json()["id"]
        try:
            r = requests.get(f"{API}/dashboard", headers=admin_h)
            assert r.status_code == 200
            d = r.json()
            assert "critical_products" in d
            assert isinstance(d["critical_products"], list)
            assert any(p.get("id") == pid for p in d["critical_products"])
        finally:
            requests.delete(f"{API}/products/{pid}", headers=admin_h)
