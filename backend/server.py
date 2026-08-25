from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# Türkçe kategori metinlerini büyük/küçük harf ve aksan farklarından bağımsız karşılaştır.
def _category_key(value):
    text = unicodedata.normalize("NFKD", str(value or "").casefold())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.translate(str.maketrans({"ı": "i"}))
    text = "".join(ch if ch.isalnum() else " " for ch in text)
    return " ".join(text.split())


import os
import io
import uuid
import base64
import mimetypes
import logging
import unicodedata
from datetime import datetime, timezone, timedelta, date
from typing import Optional, List
from zoneinfo import ZoneInfo

import bcrypt
import jwt
import httpx
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

# ---------- Setup ----------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="CNC Takımhane API")
# Hurda / Kullanım Dışı modülü sürüm işareti: 2026-08-21
SCRAP_MODULE_VERSION = "2026-08-21"
api = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]
EMAIL_BASE_URL = "https://integrations.emergentagent.com"
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "CNC Takımhane")
OWNER_EMAIL = os.environ.get("OWNER_EMAIL", "")

logger = logging.getLogger("cnc")
logging.basicConfig(level=logging.INFO)


def now_utc():
    return datetime.now(timezone.utc)


def new_id() -> str:
    return str(uuid.uuid4())


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "type": "access",
               "exp": now_utc() + timedelta(hours=12)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def set_auth_cookie(response: Response, token: str):
    response.set_cookie("access_token", token, httponly=True, secure=True,
                        samesite="none", max_age=12 * 3600, path="/")


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Yetkisiz erişim")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Geçersiz token")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
        # role is re-read from DB every request (skill: re-derive privileged authz from source of truth)
        user["role"] = user.get("role") or "viewer"
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Oturum süresi doldu")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Geçersiz token")


async def require_admin(user=Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yönetici yetkisi gerekli. Sadece görüntüleme yapabilirsiniz.")
    return user


def clean(doc: dict) -> dict:
    if doc and "_id" in doc:
        doc.pop("_id")
    return doc


# ---------- Models ----------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class ForgotPasswordIn(BaseModel):
    email: EmailStr


class ResetPasswordIn(BaseModel):
    token: str
    new_password: str = Field(min_length=6)


class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=6)


class ProductIn(BaseModel):
    code: Optional[str] = ""
    name: str
    category: str
    unit: str = "adet"
    min_stock: float = 0.0
    current_stock: float = 0.0
    location: Optional[str] = ""
    quality: Optional[str] = ""
    brand: Optional[str] = ""
    is_special: bool = False
    image_url: Optional[str] = ""


class QuickStockAdjustmentIn(BaseModel):
    quantity: float
    direction: str
    note: Optional[str] = ""


class ToolHolderIn_Model(BaseModel):
    name: str
    brand: Optional[str] = ""
    cutting_tool_code_name: Optional[str] = ""
    type: Optional[str] = ""
    length: Optional[str] = ""
    diameter: Optional[str] = ""
    min_stock: float = 0.0
    current_stock: float = 0.0
    location: Optional[str] = ""
    note: Optional[str] = ""
    image_url: Optional[str] = ""


class ToolHolderStockIn(BaseModel):
    quantity: float
    supplier: Optional[str] = ""
    supplier_id: Optional[str] = None
    note: Optional[str] = ""


class ToolHolderStockOut(BaseModel):
    quantity: float
    machine_id: str
    personnel_id: Optional[str] = None
    note: Optional[str] = ""


class ToolHolderScrapIn(BaseModel):
    quantity: float
    scrap_reason: str
    description: Optional[str] = ""
    scrap_date: str
    location: Optional[str] = ""
    operator_personnel_id: Optional[str] = ""
    approver_personnel_id: Optional[str] = ""
    witness_personnel_id: Optional[str] = ""
    approved_by: Optional[str] = ""
    witness: Optional[str] = ""
    before_image_url: Optional[str] = ""
    after_image_url: Optional[str] = ""


class PersonnelIn(BaseModel):
    first_name: str
    last_name: str
    department: Optional[str] = ""
    image_url: Optional[str] = ""


class MachineIn(BaseModel):
    code: str
    name: str
    brand: Optional[str] = ""
    model: Optional[str] = ""
    type: Optional[str] = ""
    description: Optional[str] = ""


class StockInIn(BaseModel):
    product_id: str
    quantity: float
    supplier: Optional[str] = ""
    supplier_id: Optional[str] = None
    note: Optional[str] = ""
    transaction_date: Optional[str] = ""


class StockOutIn(BaseModel):
    product_id: str
    quantity: float
    personnel_id: str
    machine_id: str
    toolholder_id: Optional[str] = ""
    note: Optional[str] = ""
    transaction_date: Optional[str] = ""
    production_product: Optional[str] = ""


class SharpeningOutIn(BaseModel):
    product_id: str
    quantity: float
    helix_length: Optional[str] = ""
    diameter: Optional[str] = ""
    full_length: Optional[str] = ""
    process_type: str
    company: str
    sent_date: str
    note: Optional[str] = ""


class SharpeningInIn(BaseModel):
    record_id: str
    company: str
    waybill_number: str
    received_date: str
    quantity: Optional[float] = None
    target_product_id: Optional[str] = None
    return_helix_length: Optional[str] = ""
    return_diameter: Optional[str] = ""
    return_full_length: Optional[str] = ""
    note: Optional[str] = ""


class SharpeningUpdateIn(BaseModel):
    quantity: float
    helix_length: str = ""
    diameter: str = ""
    full_length: str = ""
    process_type: str
    company: str
    sent_date: str
    note: str = ""
    returned_quantity: float = 0
    return_company: str = ""
    waybill_number: str = ""
    received_date: str = ""
    return_note: str = ""
    return_product_id: Optional[str] = None
    return_helix_length: str = ""
    return_diameter: str = ""
    return_full_length: str = ""


class ToolTrialIn(BaseModel):
    comparison_name: str = ""
    trial_date: str
    part_name: str
    material: Optional[str] = ""
    hardness: Optional[str] = ""
    machine_id: Optional[str] = None
    operation_type: str = "Tornalama"
    product_id: Optional[str] = None
    product_brand: Optional[str] = ""
    product_code: Optional[str] = ""
    insert_grade: Optional[str] = ""
    tool_diameter: Optional[str] = ""
    quantity_used: float = 0
    price: float = 0
    spindle_speed: Optional[str] = ""
    cutting_speed: Optional[str] = ""
    feed_rate: Optional[str] = ""
    depth_of_cut: Optional[str] = ""
    drill_diameter: Optional[str] = ""
    feed_per_tooth: Optional[str] = ""
    runtime_minutes: float = 0
    parts_machined: float = 0
    wear_result: Optional[str] = ""
    result: Optional[str] = ""
    technical_comment: Optional[str] = ""
    coolant: Optional[str] = ""


class RecipeLineIn(BaseModel):
    kind: str
    reference_id: str
    quantity: float
    note: Optional[str] = ""


class RecipeIn(BaseModel):
    name: str
    workpiece_name: str
    note: Optional[str] = ""
    lines: List[RecipeLineIn]


class RecipeBindIn(BaseModel):
    machine_id: str


class SupplierIn(BaseModel):
    name: str
    contact_person: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""
    note: Optional[str] = ""


class OrderItemIn(BaseModel):
    product_id: Optional[str] = None
    product_code: Optional[str] = ""
    product_name: Optional[str] = ""
    category: Optional[str] = "Diğer"
    unit: Optional[str] = "adet"
    quantity: float


class OrderIn(BaseModel):
    supplier_id: str
    delivery_date: Optional[str] = None  # ISO date "YYYY-MM-DD"
    note: Optional[str] = ""
    items: List[OrderItemIn]


class ReceiveItemIn(BaseModel):
    product_id: str
    quantity: float


class ReceiveIn(BaseModel):
    items: List[ReceiveItemIn]


# ---------- Sample Data ----------
SAMPLE_PRODUCTS = [
    {"code": "YZK00001", "name": "Kesici Uç CNMG 120408", "category": "Kesici Uç", "unit": "adet", "min_stock": 20, "current_stock": 45, "location": "Raf A-1", "quality": "TiAlN", "brand": "Sandvik"},
    {"code": "YZK00002", "name": "Kesici Uç DNMG 150608", "category": "Kesici Uç", "unit": "adet", "min_stock": 15, "current_stock": 8, "location": "Raf A-2", "quality": "TiN", "brand": "Kennametal"},
    {"code": "YZK00003", "name": "HSS Matkap Ucu Ø8mm", "category": "Matkap", "unit": "adet", "min_stock": 25, "current_stock": 60, "location": "Raf B-1", "quality": "HSS", "brand": "Bosch"},
    {"code": "YZK00004", "name": "Karbür Matkap Ucu Ø10mm", "category": "Matkap", "unit": "adet", "min_stock": 10, "current_stock": 4, "location": "Raf B-2", "quality": "Karbür", "brand": "Guhring"},
    {"code": "YZK00005", "name": "Kater DCLNR 2525M-12", "category": "Kater", "unit": "adet", "min_stock": 3, "current_stock": 7, "location": "Dolap C", "quality": "", "brand": "Iscar"},
    {"code": "YZK00006", "name": "ER32 Pens Seti", "category": "Apparat", "unit": "set", "min_stock": 2, "current_stock": 3, "location": "Dolap D", "quality": "", "brand": "Regofix"},
    {"code": "YZK00007", "name": "Kumpas 0-150mm Dijital", "category": "Ölçüm Aleti", "unit": "adet", "min_stock": 5, "current_stock": 12, "location": "Dolap E", "quality": "", "brand": "Mitutoyo"},
    {"code": "YZK00008", "name": "Mikrometre 0-25mm", "category": "Ölçüm Aleti", "unit": "adet", "min_stock": 4, "current_stock": 2, "location": "Dolap E", "quality": "", "brand": "Mitutoyo"},
]

SAMPLE_PERSONNEL = [
    {"first_name": "Ahmet", "last_name": "Yılmaz", "department": "CNC Tornacı"},
    {"first_name": "Mehmet", "last_name": "Kaya", "department": "CNC Dik İşlemeci"},
    {"first_name": "Ali", "last_name": "Demir", "department": "CNC Tornacı"},
    {"first_name": "Ayşe", "last_name": "Öztürk", "department": "Üretim Mühendisi"},
    {"first_name": "Mustafa", "last_name": "Şahin", "department": "Taşlamacı"},
]

SAMPLE_MACHINES = [
    {"code": "T-01", "name": "Torna 01", "brand": "Mazak", "model": "Quick Turn 250", "type": "CNC Torna", "description": ""},
    {"code": "T-02", "name": "Torna 02", "brand": "Doosan", "model": "Puma 2600", "type": "CNC Torna", "description": ""},
    {"code": "F-01", "name": "Freze 01", "brand": "Haas", "model": "VF-2", "type": "CNC Freze / Dik İşleme", "description": "3 Eksen"},
    {"code": "F-02", "name": "Freze 02", "brand": "DMG Mori", "model": "DMU 50", "type": "CNC Freze / Dik İşleme", "description": "5 Eksen"},
]


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.products.create_index("code", unique=True)
    # Drop obsolete unique index on personnel.reg_no if it exists (schema simplified — user removed sicil no)
    try:
        await db.personnel.drop_index("reg_no_1")
    except Exception:
        pass
    await db.machines.create_index("code", unique=True)
    await db.movements.create_index("created_at")
    await db.toolholder_movements.create_index("created_at")
    await db.sharpening_records.create_index("created_at")
    await db.sharpening_records.create_index("status")
    await db.tool_trials.create_index("trial_date")
    await db.tool_trials.create_index("comparison_name")
    await db.recipes.create_index("status")

    admin_email = os.environ.get("ADMIN_EMAIL", "").lower()
    admin_pw = os.environ.get("ADMIN_PASSWORD", "")
    admin_name = os.environ.get("ADMIN_NAME", "Admin")
    # Migration: any existing user without a role becomes 'viewer' (fail-safe: least privilege)
    await db.users.update_many({"role": {"$exists": False}}, {"$set": {"role": "viewer"}})
    if admin_email and admin_pw:
        existing = await db.users.find_one({"email": admin_email})
        if not existing:
            await db.users.insert_one({
                "id": new_id(), "email": admin_email, "name": admin_name,
                "password_hash": hash_password(admin_pw),
                "role": "admin",
                "created_at": now_utc().isoformat(),
            })
            logger.info(f"Admin seeded: {admin_email}")
        else:
            update = {"role": "admin"}
            if not verify_password(admin_pw, existing.get("password_hash", "")):
                update["password_hash"] = hash_password(admin_pw)
            await db.users.update_one({"email": admin_email}, {"$set": update})

    if await db.products.count_documents({}) == 0:
        # Guard: only auto-seed sample data if never seeded before. After wipe-all
        # (or manual seed flag insert) sample data will NOT return on restart.
        seeded_marker = await db.settings.find_one({"key": "sample_seeded"})
        if not seeded_marker:
            for p in SAMPLE_PRODUCTS:
                await db.products.insert_one({**p, "id": new_id(), "created_at": now_utc().isoformat()})
            if await db.personnel.count_documents({}) == 0:
                for p in SAMPLE_PERSONNEL:
                    await db.personnel.insert_one({**p, "id": new_id(), "created_at": now_utc().isoformat()})
            if await db.machines.count_documents({}) == 0:
                for m in SAMPLE_MACHINES:
                    await db.machines.insert_one({**m, "id": new_id(), "created_at": now_utc().isoformat()})
            await db.settings.update_one({"key": "sample_seeded"},
                                         {"$set": {"key": "sample_seeded", "value": True,
                                                   "at": now_utc().isoformat()}},
                                         upsert=True)

    # Apply the supplied common image to all existing Parmak Freze products.
    parmak_image_path = ROOT_DIR / "assets" / "parmakfreze.jpg"
    if parmak_image_path.exists():
        image_data = base64.b64encode(parmak_image_path.read_bytes()).decode("ascii")
        image_url = f"data:image/jpeg;base64,{image_data}"
        matching = await db.products.find({}).to_list(5000)
        # Kategori alanı önceliklidir; boşluk/büyük-küçük harf farklarını tolere et.
        matched_ids = []
        for product in matching:
            category = _category_key(product.get("category", ""))
            searchable = _category_key(" ".join(str(product.get(k, "")) for k in ("name", "category", "brand")))
            if "parmak freze" in category or "parmak freze" in searchable:
                matched_ids.append(product["id"])
        if matched_ids:
            await db.products.update_many({"id": {"$in": matched_ids}}, {"$set": {"image_url": image_url}})
            logger.info(f"Parmak Freze görseli uygulandı: {len(matched_ids)} ürün")

    # Karbür Matkap kategorisindeki tüm ürünlere ortak görseli uygula.
    carbide_image_path = ROOT_DIR / "assets" / "karbur.jpg"
    if carbide_image_path.exists():
        carbide_data = base64.b64encode(carbide_image_path.read_bytes()).decode("ascii")
        carbide_url = f"data:image/jpeg;base64,{carbide_data}"
        carbide_products = await db.products.find({}).to_list(5000)
        carbide_ids = [
            product["id"] for product in carbide_products
            if "karbur matkap" in _category_key(product.get("category", ""))
        ]
        if carbide_ids:
            await db.products.update_many({"id": {"$in": carbide_ids}}, {"$set": {"image_url": carbide_url}})
            logger.info(f"Karbür Matkap görseli uygulandı: {len(carbide_ids)} ürün")

    # Kılavuz kategorisindeki tüm ürünlere ortak görseli uygula.
    guide_image_path = ROOT_DIR / "assets" / "kilavuz.jpg"
    if guide_image_path.exists():
        guide_data = base64.b64encode(guide_image_path.read_bytes()).decode("ascii")
        guide_url = f"data:image/jpeg;base64,{guide_data}"
        guide_products = await db.products.find({}).to_list(5000)
        guide_ids = [
            product["id"] for product in guide_products
            if "kilavuz" in _category_key(product.get("category", ""))
        ]
        if guide_ids:
            await db.products.update_many({"id": {"$in": guide_ids}}, {"$set": {"image_url": guide_url}})
            logger.info(f"Kılavuz görseli uygulandı: {len(guide_ids)} ürün")


@app.on_event("shutdown")
async def shutdown():
    client.close()


async def send_critical_stock_email(product: dict):
    if not EMAIL_KEY or not OWNER_EMAIL:
        return
    subject = f"[Kritik Stok] {product['name']} minimum seviyenin altına düştü"
    html = f"""
    <table width="100%" style="font-family: Arial, sans-serif; background:#0f172a; color:#f8fafc; padding:24px; border-radius:12px;">
      <tr><td>
        <h2 style="color:#f87171; margin:0 0 12px 0;">Kritik Stok Uyarısı</h2>
        <p style="color:#cbd5e1; font-size:14px;">Aşağıdaki ürün minimum stok seviyesinin altına düştü.</p>
        <table style="width:100%; border-collapse:collapse; background:#1e293b; border-radius:8px; margin-top:16px;">
          <tr><td style="padding:12px; color:#94a3b8;">Ürün Kodu</td><td style="padding:12px; font-weight:bold;">{product['code']}</td></tr>
          <tr><td style="padding:12px; color:#94a3b8;">Ürün Adı</td><td style="padding:12px; font-weight:bold;">{product['name']}</td></tr>
          <tr><td style="padding:12px; color:#94a3b8;">Kategori</td><td style="padding:12px;">{product['category']}</td></tr>
          <tr><td style="padding:12px; color:#94a3b8;">Mevcut Stok</td><td style="padding:12px; color:#f87171; font-weight:bold;">{product['current_stock']} {product['unit']}</td></tr>
          <tr><td style="padding:12px; color:#94a3b8;">Minimum Stok</td><td style="padding:12px;">{product['min_stock']} {product['unit']}</td></tr>
        </table>
        <p style="color:#94a3b8; font-size:12px; margin-top:24px;">CNC Takımhane Stok Takip Sistemi</p>
      </td></tr>
    </table>
    """
    payload = {"to": [OWNER_EMAIL], "subject": subject, "html": html, "from_name": EMAIL_FROM_NAME}
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.post(f"{EMAIL_BASE_URL}/api/v1/email/send",
                             headers={"X-Email-Key": EMAIL_KEY}, json=payload)
            r.raise_for_status()
    except Exception as e:
        logger.error(f"Kritik stok e-posta hatası: {e}")


# ---------- Auth ----------
@api.post("/auth/register")
async def register(body: RegisterIn, response: Response):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Bu e-posta zaten kayıtlı")
    uid = new_id()
    # Newly registered users are viewers by default — only admin (seeded via env) can mutate data.
    await db.users.insert_one({
        "id": uid, "email": email, "name": body.name,
        "password_hash": hash_password(body.password),
        "role": "viewer",
        "created_at": now_utc().isoformat(),
    })
    token = create_access_token(uid, email)
    set_auth_cookie(response, token)
    return {"user": {"id": uid, "email": email, "name": body.name, "role": "viewer"}, "access_token": token}


@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    token = create_access_token(user["id"], email)
    set_auth_cookie(response, token)
    role = user.get("role") or "viewer"
    return {"user": {"id": user["id"], "email": email, "name": user["name"], "role": role},
            "access_token": token}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


# ---------- Password reset ----------
import secrets, hashlib


def _hash_token(t: str) -> str:
    return hashlib.sha256(t.encode()).hexdigest()


async def send_password_reset_email(to_email: str, reset_url: str, user_name: str):
    if not RESEND_API_KEY:
        logger.warning("RESEND_API_KEY not configured — cannot send password reset")
        return False
    subject = "Yazkan Döküm Takımhane — Şifre Sıfırlama"
    html = f"""
    <div style="font-family: Arial, sans-serif; max-width:560px; margin:0 auto; background:#0f172a; color:#f8fafc; padding:32px; border-radius:12px;">
      <h2 style="color:#60a5fa; margin:0 0 8px 0;">Şifre Sıfırlama</h2>
      <p style="color:#cbd5e1; font-size:14px;">Merhaba <strong>{user_name}</strong>,</p>
      <p style="color:#cbd5e1; font-size:14px;">Yazkan Döküm Takımhane sisteminde şifre sıfırlama talebinde bulundunuz. Yeni bir şifre belirlemek için aşağıdaki bağlantıya tıklayın:</p>
      <p style="margin:28px 0;">
        <a href="{reset_url}" style="display:inline-block; padding:14px 24px; background:#2563eb; color:#fff; text-decoration:none; border-radius:8px; font-weight:bold;">Şifreyi Sıfırla</a>
      </p>
      <p style="color:#94a3b8; font-size:12px;">Bu bağlantı 1 saat içinde geçerliliğini yitirir. Talep sizin tarafınızdan yapılmadıysa bu e-postayı yok sayabilirsiniz.</p>
      <p style="color:#64748b; font-size:11px; margin-top:24px; word-break:break-all;">Bağlantı çalışmıyorsa: {reset_url}</p>
    </div>
    """
    payload = {"from": f"{EMAIL_FROM_NAME} <onboarding@resend.dev>", "to": [to_email], "subject": subject, "html": html}
    try:
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.post("https://api.resend.com/emails",
                 headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
                 json=payload )
            r.raise_for_status()
        return True
    except Exception as e:
        logger.error(f"Password reset email failed: {e}")
        return False


@api.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordIn):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    # Always return the same response — don't leak whether the email exists
    if user:
        raw_token = secrets.token_urlsafe(32)
        await db.password_resets.insert_one({
            "id": new_id(),
            "user_id": user["id"],
            "email": email,
            "token_hash": _hash_token(raw_token),
            "expires_at": (now_utc() + timedelta(hours=1)).isoformat(),
            "used": False,
            "created_at": now_utc().isoformat(),
        })
        app_url = os.environ.get("APP_URL", "").rstrip("/")
        reset_url = f"{app_url}/sifre-sifirla?token={raw_token}"
        await send_password_reset_email(email, reset_url, user.get("name", email))
    return {"ok": True, "message": "Eğer bu e-posta sistemde kayıtlıysa, şifre sıfırlama bağlantısı gönderildi."}


@api.post("/auth/reset-password")
async def reset_password(body: ResetPasswordIn):
    if not body.token:
        raise HTTPException(status_code=400, detail="Token gerekli")
    token_hash = _hash_token(body.token)
    rec = await db.password_resets.find_one({"token_hash": token_hash, "used": False})
    if not rec:
        raise HTTPException(status_code=400, detail="Geçersiz veya kullanılmış bağlantı")
    try:
        expires = datetime.fromisoformat(rec["expires_at"])
    except Exception:
        raise HTTPException(status_code=400, detail="Bağlantı geçersiz")
    if now_utc() > expires:
        raise HTTPException(status_code=400, detail="Bağlantının süresi dolmuş, tekrar talep edin")
    user = await db.users.find_one({"id": rec["user_id"]})
    if not user:
        raise HTTPException(status_code=400, detail="Kullanıcı bulunamadı")
    await db.users.update_one({"id": user["id"]},
                              {"$set": {"password_hash": hash_password(body.new_password)}})
    await db.password_resets.update_one({"id": rec["id"]},
                                        {"$set": {"used": True, "used_at": now_utc().isoformat()}})
    return {"ok": True, "message": "Şifreniz güncellendi. Yeni şifrenizle giriş yapabilirsiniz."}


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


@api.post("/auth/change-password")
async def change_password(body: ChangePasswordIn, user=Depends(get_current_user)):
    """Giriş yapmış kullanıcı kendi şifresini değiştirir."""
    current = await db.users.find_one({"id": user["id"]})
    if not current or not verify_password(body.current_password, current.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Mevcut şifre hatalı")
    if body.current_password == body.new_password:
        raise HTTPException(status_code=400, detail="Yeni şifre eskisiyle aynı olamaz")
    await db.users.update_one({"id": user["id"]},
                              {"$set": {"password_hash": hash_password(body.new_password)}})
    return {"ok": True, "message": "Şifreniz güncellendi"}

    class AdminResetPasswordIn(BaseModel):
        user_id: str
        new_password: str = Field(min_length=6)

@api.get("/admin/users")
async def admin_list_users(user=Depends(require_admin)):
        return await db.users.find(
            {}, {"_id": 0, "password_hash": 0}
        ).sort("name", 1).to_list(1000)
class AdminResetPasswordIn(BaseModel):
    user_id: str
    new_password: str = Field(min_length=6)
@api.get("/admin/users")
async def admin_list_users(user=Depends(require_admin)):
    return await db.users.find(
        {}, {"_id": 0, "password_hash": 0}
    ).sort("name", 1).to_list(1000)

@api.post("/admin/users/reset-password")
async def admin_reset_user_password(body: AdminResetPasswordIn, user=Depends(require_admin)):
    target = await db.users.find_one({"id": body.user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if target.get("role") == "admin" and target.get("id") != user.get("id"):
        raise HTTPException(status_code=403, detail="Başka bir yöneticinin şifresi bu ekrandan değiştirilemez")
    await db.users.update_one(
        {"id": body.user_id},
        {"$set": {"password_hash": hash_password(body.new_password)}}
    )
    return {"ok": True, "message": "Kullanıcı şifresi güncellendi"}

# ---------- Critical (define BEFORE /products/{pid} routes) ----------
@api.get("/products/critical")
async def critical_products(user=Depends(get_current_user)):
    docs = await db.products.find({}, {"_id": 0}).to_list(2000)
    return [d for d in docs if d.get("current_stock", 0) <= d.get("min_stock", 0)]


async def _next_product_code() -> str:
    import re
    docs = await db.products.find({"code": {"$regex": r"^YZK\d+$"}}, {"code": 1}).to_list(50000)
    max_n = 0
    for d in docs:
        m = re.match(r"^YZK(\d+)$", d.get("code", ""))
        if m:
            n = int(m.group(1))
            if n > max_n:
                max_n = n
    return f"YZK{max_n + 1:05d}"


# ---------- Image upload ----------
@api.post("/images/upload")
async def upload_image(file: UploadFile = File(...), user=Depends(require_admin)):
    allowed = {"image/jpeg", "image/png", "image/webp"}
    content_type = file.content_type or mimetypes.guess_type(file.filename or "")[0] or ""
    if content_type not in allowed:
        raise HTTPException(status_code=400, detail="Sadece JPG, PNG veya WEBP görsel yükleyebilirsiniz")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Görsel boyutu en fazla 5 MB olabilir")
    encoded = base64.b64encode(content).decode("ascii")
    return {"image_url": f"data:{content_type};base64,{encoded}", "filename": file.filename or "gorsel"}


# ---------- Products ----------
@api.post("/products/apply-category-image")
async def apply_category_image(body: dict, user=Depends(require_admin)):
    category_key = _category_key(body.get("category", ""))
    asset_map = {
        "kilavuz": ("kilavuz", "kilavuz.jpg"),
        "t line matkap": ("t line matkap", "tline.jpg"),
    }
    if category_key not in asset_map:
        raise HTTPException(status_code=400, detail="Desteklenmeyen kategori görseli")
    category_match, asset_name = asset_map[category_key]
    image_path = ROOT_DIR / "assets" / asset_name
    if not image_path.exists():
        raise HTTPException(status_code=500, detail="Kılavuz görseli sunucuda bulunamadı")
    encoded = base64.b64encode(image_path.read_bytes()).decode("ascii")
    image_url = f"data:image/jpeg;base64,{encoded}"
    products = await db.products.find({}, {"id": 1, "category": 1, "image_url": 1}).to_list(5000)
    ids = [
        p.get("id") for p in products
        if p.get("id")
        and category_match in _category_key(p.get("category", ""))
        and not p.get("image_url")
    ]
    if ids:
        result = await db.products.update_many({"id": {"$in": ids}}, {"$set": {"image_url": image_url}})
        return {"ok": True, "updated": result.modified_count, "matched": len(ids)}
    return {"ok": True, "updated": 0, "matched": 0}


@api.get("/products")
async def list_products(user=Depends(get_current_user)):
    # Görsel kategori migration'ları startup sırasında bir kez çalışır.
    # Listeleme endpoint'inde tekrar tekrar dosya okuyup toplu update yapmak,
    # her sayfa açılışını gereksiz yere yavaşlatır; bu sıcak yolda yalnızca okuma yapıyoruz.
    products = await db.products.find({}, {"_id": 0}).sort("code", 1).to_list(2000)
    sharpening_totals = {}
    open_records = await db.sharpening_records.find({"status": {"$in": ["sent", "partial"]}}, {"_id": 0, "product_id": 1, "quantity": 1, "returned_quantity": 1, "remaining_quantity": 1}).to_list(5000)
    for record in open_records:
        remaining = record.get("remaining_quantity")
        if remaining is None:
            remaining = float(record.get("quantity", 0)) - float(record.get("returned_quantity", 0))
        if float(remaining or 0) > 0:
            product_id = record.get("product_id")
            sharpening_totals[product_id] = sharpening_totals.get(product_id, 0) + float(remaining)
    for product in products:
        product["in_sharpening"] = round(sharpening_totals.get(product.get("id"), 0), 6)
    return products


@api.post("/products")
async def create_product(body: ProductIn, user=Depends(require_admin)):
    data = body.model_dump()
    code = (data.get("code") or "").strip()
    if not code:
        code = await _next_product_code()
    if await db.products.find_one({"code": code}):
        raise HTTPException(status_code=400, detail="Bu ürün kodu zaten mevcut")
    data["code"] = code
    doc = {**data, "id": new_id(), "created_at": now_utc().isoformat()}
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/products/{pid}")
async def update_product(pid: str, body: ProductIn, user=Depends(require_admin)):
    existing = await db.products.find_one({"id": pid})
    if not existing:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    data = body.model_dump()
    code = (data.get("code") or existing["code"]).strip()
    if code != existing["code"] and await db.products.find_one({"code": code}):
        raise HTTPException(status_code=400, detail="Bu ürün kodu zaten mevcut")
    data["code"] = code
    await db.products.update_one({"id": pid}, {"$set": data})
    return await db.products.find_one({"id": pid}, {"_id": 0})


@api.post("/products/{pid}/quick-stock")
async def quick_stock_adjustment(pid: str, body: QuickStockAdjustmentIn, user=Depends(require_admin)):
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    if body.direction not in {"in", "out"}:
        raise HTTPException(status_code=400, detail="Geçersiz stok yönü")
    product = await db.products.find_one({"id": pid})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    current = float(product.get("current_stock", 0))
    delta = body.quantity if body.direction == "in" else -body.quantity
    new_stock = round(current + delta, 6)
    if new_stock < 0:
        raise HTTPException(status_code=400, detail="Stok eksiye düşemez")
    movement = {
        "id": new_id(), "type": "quick_in" if body.direction == "in" else "quick_out",
        "product_id": product["id"], "product_code": product.get("code", ""),
        "product_name": product.get("name", ""), "quantity": body.quantity,
        "unit_price": 0, "total": 0, "note": body.note or "Ürünler sayfası hızlı stok işlemi",
        "user_id": user["id"], "user_name": user["name"], "created_at": now_utc().isoformat(),
    }
    await db.products.update_one({"id": pid}, {"$set": {"current_stock": new_stock}})
    await db.movements.insert_one(movement)
    return {"ok": True, "new_stock": new_stock, "movement": clean(movement)}


@api.delete("/products/{pid}")
async def delete_product(pid: str, user=Depends(require_admin)):
    r = await db.products.delete_one({"id": pid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return {"ok": True}


# ---------- Personnel ----------
@api.get("/personnel")
async def list_personnel(user=Depends(get_current_user)):
    return await db.personnel.find({}, {"_id": 0}).sort("first_name", 1).to_list(2000)


@api.post("/personnel")
async def create_personnel(body: PersonnelIn, user=Depends(require_admin)):
    doc = {**body.model_dump(), "id": new_id(), "created_at": now_utc().isoformat()}
    await db.personnel.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/personnel/{pid}")
async def update_personnel(pid: str, body: PersonnelIn, user=Depends(require_admin)):
    existing = await db.personnel.find_one({"id": pid})
    if not existing:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    await db.personnel.update_one({"id": pid}, {"$set": body.model_dump()})
    return await db.personnel.find_one({"id": pid}, {"_id": 0})


@api.delete("/personnel/{pid}")
async def delete_personnel(pid: str, user=Depends(require_admin)):
    r = await db.personnel.delete_one({"id": pid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    return {"ok": True}


@api.get("/personnel/import/template")
async def personnel_import_template(user=Depends(get_current_user)):
    wb = Workbook(); ws = wb.active; ws.title = "Personel"
    ws.append(["first_name", "last_name", "department", "image_url"])
    ws.append(["Örnek", "Personel", "CNC Tornacı", ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="personel_sablonu.xlsx"'})


@api.post("/personnel/import")
async def personnel_import(file: UploadFile = File(...), commit: bool = False, user=Depends(require_admin)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyası yükleyin")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")
    rows = list(wb.active.iter_rows(values_only=True))
    if len(rows) < 2: raise HTTPException(status_code=400, detail="Dosyada veri yok")
    header = [str(c).strip().lower() if c else "" for c in rows[0]]; idx = {h: i for i, h in enumerate(header) if h}
    if "first_name" not in idx or "last_name" not in idx: raise HTTPException(status_code=400, detail="first_name ve last_name zorunludur")
    def get(row, key):
        i = idx.get(key); return "" if i is None or i >= len(row) or row[i] is None else str(row[i]).strip()
    preview = []
    for row_num, row in enumerate(rows[1:], 2):
        first, last = get(row, "first_name"), get(row, "last_name")
        if not first or not last: preview.append({"row": row_num, "action": "skip", "error": "Ad ve soyad zorunludur"}); continue
        d = {"first_name": first, "last_name": last, "department": get(row, "department"), "image_url": get(row, "image_url")}
        existing = await db.personnel.find_one({"first_name": first, "last_name": last})
        preview.append({"row": row_num, "action": "update" if existing else "create", "data": d})
    if not commit: return {"committed": False, "preview": preview}
    created = updated = 0
    for item in preview:
        if item["action"] == "skip": continue
        d = item["data"]
        if item["action"] == "create": await db.personnel.insert_one({**d, "id": new_id(), "created_at": now_utc().isoformat()}); created += 1
        else: await db.personnel.update_one({"first_name": d["first_name"], "last_name": d["last_name"]}, {"$set": d}); updated += 1
    return {"committed": True, "created": created, "updated": updated, "preview": preview}


# ---------- Machines ----------
@api.get("/machines")
async def list_machines(user=Depends(get_current_user)):
    return await db.machines.find({}, {"_id": 0}).sort("code", 1).to_list(2000)


@api.post("/machines")
async def create_machine(body: MachineIn, user=Depends(require_admin)):
    if await db.machines.find_one({"code": body.code}):
        raise HTTPException(status_code=400, detail="Bu tezgah kodu zaten mevcut")
    doc = {**body.model_dump(), "id": new_id(), "created_at": now_utc().isoformat()}
    await db.machines.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/machines/{mid}")
async def update_machine(mid: str, body: MachineIn, user=Depends(require_admin)):
    existing = await db.machines.find_one({"id": mid})
    if not existing:
        raise HTTPException(status_code=404, detail="Tezgah bulunamadı")
    if body.code != existing["code"] and await db.machines.find_one({"code": body.code}):
        raise HTTPException(status_code=400, detail="Bu tezgah kodu zaten mevcut")
    await db.machines.update_one({"id": mid}, {"$set": body.model_dump()})
    return await db.machines.find_one({"id": mid}, {"_id": 0})


@api.delete("/machines/{mid}")
async def delete_machine(mid: str, user=Depends(require_admin)):
    r = await db.machines.delete_one({"id": mid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tezgah bulunamadı")
    return {"ok": True}


# ---------- Stock movements ----------
@api.post("/stock/in")
async def stock_in(body: StockInIn, user=Depends(require_admin)):
    product = await db.products.find_one({"id": body.product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    new_stock = product["current_stock"] + body.quantity
    await db.products.update_one({"id": body.product_id}, {"$set": {"current_stock": new_stock}})
    movement = {
        "id": new_id(), "type": "in", "product_id": product["id"],
        "product_code": product["code"], "product_name": product["name"],
        "quantity": body.quantity, "unit_price": 0, "total": 0,
        "supplier": body.supplier or "", "supplier_id": body.supplier_id or "", "note": body.note or "",
        "user_id": user["id"], "user_name": user["name"],
        "created_at": now_utc().isoformat(),
        "transaction_date": body.transaction_date or now_utc().date().isoformat(),
    }
    await db.movements.insert_one(movement)
    return {"ok": True, "new_stock": new_stock, "movement": clean(movement)}


@api.post("/stock/out")
async def stock_out(body: StockOutIn, user=Depends(require_admin)):
    product = await db.products.find_one({"id": body.product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    if body.quantity > product["current_stock"]:
        raise HTTPException(status_code=400, detail="Yetersiz stok")
    personnel = await db.personnel.find_one({"id": body.personnel_id})
    if not personnel:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    machine = await db.machines.find_one({"id": body.machine_id})
    if not machine:
        raise HTTPException(status_code=404, detail="Tezgah bulunamadı")
    toolholder = None
    if body.toolholder_id:
        toolholder = await db.toolholders.find_one({"id": body.toolholder_id})
        if not toolholder:
            raise HTTPException(status_code=404, detail="Takım tutucu bulunamadı")
    new_stock = product["current_stock"] - body.quantity
    await db.products.update_one({"id": body.product_id}, {"$set": {"current_stock": new_stock}})
    movement = {
        "id": new_id(), "type": "out", "product_id": product["id"],
        "product_code": product["code"], "product_name": product["name"],
        "quantity": body.quantity, "unit_price": 0, "total": 0,
        "personnel_id": personnel["id"],
        "personnel_name": f"{personnel['first_name']} {personnel['last_name']}",
        "machine_id": machine["id"], "machine_code": machine.get("code", ""), "machine_name": machine["name"],
        "toolholder_id": toolholder.get("id", "") if toolholder else "",
        "toolholder_code": toolholder.get("code", "") if toolholder else "",
        "toolholder_name": toolholder.get("name", "") if toolholder else "",
        "toolholder_brand": toolholder.get("brand", "") if toolholder else "",
        "note": body.note or "",
        "production_product": body.production_product or "",
        "user_id": user["id"], "user_name": user["name"],
        "created_at": now_utc().isoformat(),
        "transaction_date": body.transaction_date or now_utc().date().isoformat(),
    }
    await db.movements.insert_one(movement)
    critical = new_stock <= product["min_stock"]
    # Kritik stok e-postası kullanıcı isteği üzerine devre dışı — bilgi sadece Ana Panel'de görünür.
    return {"ok": True, "new_stock": new_stock, "movement": clean(movement), "critical": critical}


@api.get("/movements")
async def list_movements(
    user=Depends(get_current_user),
    type: Optional[str] = None,
    product_id: Optional[str] = None,
    personnel_id: Optional[str] = None,
    machine_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 500,
):
    q = {}
    if type:
        q["type"] = type
    if product_id:
        q["product_id"] = product_id
    if personnel_id:
        q["personnel_id"] = personnel_id
    if machine_id:
        q["machine_id"] = machine_id
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        q["$or"] = [
            {"transaction_date": date_query},
            {"transaction_date": {"$exists": False}, "created_at": {"$gte": date_from or "", "$lte": (date_to + "T23:59:59") if date_to else "9999-12-31T23:59:59"}},
        ]
    movements = await db.movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)

    # Bileme kayıtları movements koleksiyonunda tutulmadığından, Hareketler
    # ekranında ayrı ve tekil hareket satırları olarak gösterilir.
    # Normal stok çıkışıyla mükerrer oluşmaması için sharpening_record_id
    # bulunan eski hareketler tekrar eklenmez.
    sharpening_q = {}
    if product_id:
        sharpening_q["product_id"] = product_id
    if personnel_id or machine_id:
        sharpening_q = None  # Bileme kayıtlarında bu alanlar bulunmaz.
    if sharpening_q is not None:
        sharpening_records = await db.sharpening_records.find(sharpening_q, {"_id": 0}).sort("created_at", -1).to_list(5000)
        existing_sharpening_ids = {m.get("sharpening_record_id") for m in movements if m.get("sharpening_record_id")}
        for record in sharpening_records:
            record_id = record.get("id")
            if record_id in existing_sharpening_ids:
                continue
            sent_date = record.get("sent_date") or record.get("created_at", "")[:10]
            received_date = record.get("received_date") or record.get("updated_at", "")[:10]
            sent_in_range = (not date_from or (sent_date or "") >= date_from) and (not date_to or (sent_date or "") <= date_to)
            received_in_range = (not date_from or (received_date or "") >= date_from) and (not date_to or (received_date or "") <= date_to)
            if type in (None, "", "out") and sent_in_range:
                movements.append({
                    "id": f"sharpening-out-{record_id}", "type": "out", "movement_category": "Bilemeye Giden",
                    "movement_purpose": "Bilemeye gitti", "sharpening_record_id": record_id,
                    "product_id": record.get("product_id"), "product_code": record.get("product_code", ""),
                    "product_name": record.get("product_name", ""), "quantity": record.get("quantity", 0),
                    "transaction_date": sent_date, "created_at": record.get("created_at", ""),
                    "supplier": record.get("company", ""), "note": record.get("note", ""),
                    "user_name": record.get("sent_by", ""), "process_type": record.get("process_type", ""),
                    "helix_length": record.get("helix_length", ""), "diameter": record.get("diameter", ""),
                    "full_length": record.get("full_length", ""),
                })
            returned_quantity = float(record.get("returned_quantity", 0) or 0)
            if returned_quantity > 0 and (type in (None, "", "in")) and received_in_range:
                movements.append({
                    "id": f"sharpening-in-{record_id}", "type": "in", "movement_category": "Bilemeden Gelen",
                    "movement_purpose": "Bilemeden geldi", "sharpening_record_id": record_id,
                    "product_id": record.get("return_product_id") or record.get("product_id"),
                    "product_code": record.get("return_product_code") or record.get("product_code", ""),
                    "product_name": record.get("return_product_name") or record.get("product_name", ""),
                    "quantity": returned_quantity, "transaction_date": received_date,
                    "created_at": record.get("updated_at") or record.get("created_at", ""),
                    "supplier": record.get("return_company", ""),
                    "note": f"İrsaliye: {record.get('waybill_number', '')}".strip(),
                    "user_name": record.get("received_by", ""), "process_type": record.get("process_type", ""),
                    "helix_length": record.get("return_helix_length", ""),
                    "diameter": record.get("return_diameter", ""), "full_length": record.get("return_full_length", ""),
                })
        movements.sort(key=lambda m: m.get("created_at") or m.get("transaction_date") or "", reverse=True)
    return movements[:limit]


# ---------- Sharpening / Bileme ----------
@api.get("/sharpening/records")
async def list_sharpening_records(user=Depends(get_current_user), status: Optional[str] = None):
    q = {}
    if status in {"sent", "partial", "returned"}:
        q["status"] = status
    return await db.sharpening_records.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)


@api.post("/sharpening/out")
async def sharpening_out(body: SharpeningOutIn, user=Depends(require_admin)):
    product = await db.products.find_one({"id": body.product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    if body.quantity > product.get("current_stock", 0):
        raise HTTPException(status_code=400, detail="Yetersiz stok")
    if body.process_type not in {"alın bileme", "tam bileme"}:
        raise HTTPException(status_code=400, detail="Geçersiz bileme işlemi")
    company = body.company.strip()
    if not company:
        raise HTTPException(status_code=400, detail="Bileme firması zorunludur")
    new_stock = product.get("current_stock", 0) - body.quantity
    record = {
        "id": new_id(), "status": "sent", "product_id": product["id"],
        "product_code": product.get("code", ""), "product_name": product.get("name", ""),
        "unit": product.get("unit", "adet"), "quantity": body.quantity,
        "remaining_quantity": body.quantity, "returned_quantity": 0,
        "helix_length": body.helix_length or "", "diameter": body.diameter or "",
        "full_length": body.full_length or "", "process_type": body.process_type,
        "company": company, "sent_date": body.sent_date, "note": body.note or "",
        "sent_by": user.get("name", ""), "created_at": now_utc().isoformat(),
    }
    await db.products.update_one({"id": product["id"]}, {"$set": {"current_stock": new_stock}})
    await db.sharpening_records.insert_one(record)
    return {"ok": True, "new_stock": new_stock, "record": clean(record)}


@api.post("/sharpening/in")
async def sharpening_in(body: SharpeningInIn, user=Depends(require_admin)):
    record = await db.sharpening_records.find_one({"id": body.record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Bileme kaydı bulunamadı")
    remaining = float(record.get("remaining_quantity", record.get("quantity", 0)))
    if record.get("status") == "returned" or remaining <= 0:
        raise HTTPException(status_code=400, detail="Bu kayıt tamamen stoğa alınmış")
    if not body.company.strip() or not body.waybill_number.strip():
        raise HTTPException(status_code=400, detail="Firma ve irsaliye numarası zorunludur")
    qty = remaining if body.quantity is None else body.quantity
    if qty <= 0 or qty > remaining:
        raise HTTPException(status_code=400, detail="Gelen miktar kalan miktar aralığında olmalıdır")
    return_product_id = body.target_product_id or record.get("return_product_id") or record["product_id"]
    return_product = await db.products.find_one({"id": return_product_id})
    if not return_product:
        raise HTTPException(status_code=404, detail="Stoğa alınacak ürün kartı bulunamadı")
    previous_return_product_id = record.get("return_product_id")
    if previous_return_product_id and previous_return_product_id != return_product_id and float(record.get("returned_quantity", 0)) > 0:
        raise HTTPException(status_code=400, detail="Daha önce stoğa alınmış kaydın ürün kartı değiştirilemez")
    left = round(remaining - qty, 6)
    new_status = "returned" if left <= 0 else "partial"
    new_stock = round(float(return_product.get("current_stock", 0)) + qty, 6)
    update = {
        "remaining_quantity": left, "returned_quantity": float(record.get("returned_quantity", 0)) + qty,
        "status": new_status, "return_product_id": return_product_id,
        "return_product_code": return_product.get("code", ""), "return_product_name": return_product.get("name", ""),
        "return_helix_length": body.return_helix_length or record.get("return_helix_length", "") or record.get("helix_length", ""),
        "return_diameter": body.return_diameter or record.get("return_diameter", "") or record.get("diameter", ""),
        "return_full_length": body.return_full_length or record.get("return_full_length", "") or record.get("full_length", ""),
        "return_company": body.company.strip(), "waybill_number": body.waybill_number.strip(), "received_date": body.received_date,
        "received_by": user.get("name", ""), "return_note": body.note or "", "updated_at": now_utc().isoformat(),
    }
    await db.products.update_one({"id": return_product_id}, {"$set": {"current_stock": new_stock}})
    await db.sharpening_records.update_one({"id": record["id"]}, {"$set": update})
    updated = await db.sharpening_records.find_one({"id": record["id"]}, {"_id": 0})
    return {"ok": True, "new_stock": new_stock, "record": updated}


@api.put("/sharpening/records/{record_id}")
async def update_sharpening_record(record_id: str, body: SharpeningUpdateIn, user=Depends(require_admin)):
    record = await db.sharpening_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Bileme kaydı bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    if body.process_type not in {"alın bileme", "tam bileme"}:
        raise HTTPException(status_code=400, detail="Geçersiz bileme işlemi")
    if not body.company.strip():
        raise HTTPException(status_code=400, detail="Bileme firması zorunludur")
    returned = float(body.returned_quantity or 0)
    if returned < 0 or returned > body.quantity:
        raise HTTPException(status_code=400, detail="Stoğa dönen miktar, gönderilen miktar aralığında olmalıdır")
    product = await db.products.find_one({"id": record["product_id"]})
    if not product:
        raise HTTPException(status_code=404, detail="Bağlı ürün bulunamadı")
    old_quantity = float(record.get("quantity", 0))
    old_returned = float(record.get("returned_quantity", 0))
    old_return_product_id = record.get("return_product_id") or record["product_id"]
    new_return_product_id = body.return_product_id or old_return_product_id
    return_product = await db.products.find_one({"id": new_return_product_id})
    if not return_product:
        raise HTTPException(status_code=404, detail="Stoğa alınacak ürün kartı bulunamadı")
    if old_returned > 0 and new_return_product_id != old_return_product_id:
        raise HTTPException(status_code=400, detail="Daha önce stoğa alınmış kaydın ürün kartı değiştirilemez")
    original_stock = round(float(product.get("current_stock", 0)) - (body.quantity - old_quantity), 6)
    if original_stock < 0:
        raise HTTPException(status_code=400, detail="Bu düzeltme stok miktarını eksiye düşürüyor")
    if new_return_product_id == record["product_id"]:
        original_stock = round(original_stock + returned - old_returned, 6)
        if original_stock < 0:
            raise HTTPException(status_code=400, detail="Bu düzeltme stok miktarını eksiye düşürüyor")
        await db.products.update_one({"id": product["id"]}, {"$set": {"current_stock": original_stock}})
    else:
        await db.products.update_one({"id": product["id"]}, {"$set": {"current_stock": original_stock}})
        target_stock = round(float(return_product.get("current_stock", 0)) + returned, 6)
        await db.products.update_one({"id": return_product["id"]}, {"$set": {"current_stock": target_stock}})
    remaining = round(body.quantity - returned, 6)
    status = "returned" if remaining <= 0 else ("partial" if returned > 0 else "sent")
    update = {
        "quantity": body.quantity, "remaining_quantity": remaining, "returned_quantity": returned,
        "helix_length": body.helix_length or "", "diameter": body.diameter or "", "full_length": body.full_length or "",
        "process_type": body.process_type, "company": body.company.strip(), "sent_date": body.sent_date,
        "note": body.note or "", "status": status, "return_product_id": new_return_product_id,
        "return_product_code": return_product.get("code", ""), "return_product_name": return_product.get("name", ""),
        "return_helix_length": body.return_helix_length or record.get("return_helix_length", "") or body.helix_length or "",
        "return_diameter": body.return_diameter or record.get("return_diameter", "") or body.diameter or "",
        "return_full_length": body.return_full_length or record.get("return_full_length", "") or body.full_length or "",
        "return_company": body.return_company.strip(), "waybill_number": body.waybill_number.strip(), "received_date": body.received_date,
        "return_note": body.return_note or "", "updated_at": now_utc().isoformat(),
    }
    await db.sharpening_records.update_one({"id": record_id}, {"$set": update})
    updated = await db.sharpening_records.find_one({"id": record_id}, {"_id": 0})
    return {"ok": True, "new_stock": original_stock, "record": updated}


@api.delete("/sharpening/records/{record_id}")
async def delete_sharpening_record(record_id: str, user=Depends(require_admin)):
    record = await db.sharpening_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Bileme kaydı bulunamadı")
    product = await db.products.find_one({"id": record["product_id"]})
    if not product:
        raise HTTPException(status_code=404, detail="Bağlı ürün bulunamadı")
    returned = float(record.get("returned_quantity", 0))
    restore = float(record.get("quantity", 0)) - returned
    return_product_id = record.get("return_product_id") or record["product_id"]
    if return_product_id == record["product_id"]:
        new_stock = round(float(product.get("current_stock", 0)) + restore, 6)
        if new_stock < 0:
            raise HTTPException(status_code=400, detail="Kayıt silinemiyor; stok dengesi eksiye düşüyor")
        await db.products.update_one({"id": product["id"]}, {"$set": {"current_stock": new_stock}})
    else:
        new_original_stock = round(float(product.get("current_stock", 0)) + restore, 6)
        return_product = await db.products.find_one({"id": return_product_id})
        if not return_product:
            raise HTTPException(status_code=404, detail="Stoğa alınan ürün kartı bulunamadı")
        new_return_stock = round(float(return_product.get("current_stock", 0)) - returned, 6)
        if new_return_stock < 0:
            raise HTTPException(status_code=400, detail="Kayıt silinemiyor; stoğa alınan kartta yeterli miktar yok")
        await db.products.update_one({"id": product["id"]}, {"$set": {"current_stock": new_original_stock}})
        await db.products.update_one({"id": return_product_id}, {"$set": {"current_stock": new_return_stock}})
        new_stock = new_original_stock
    await db.sharpening_records.delete_one({"id": record_id})
    return {"ok": True, "new_stock": new_stock}


# ---------- Tool trials / Kesici takım denemeleri ----------
@api.get("/tool-trials")
async def list_tool_trials(user=Depends(get_current_user), comparison_name: Optional[str] = None):
    q = {"comparison_name": comparison_name} if comparison_name else {}
    return await db.tool_trials.find(q, {"_id": 0}).sort("trial_date", -1).to_list(5000)


@api.post("/tool-trials")
async def create_tool_trial(body: ToolTrialIn, user=Depends(require_admin)):
    if not body.part_name.strip():
        raise HTTPException(status_code=400, detail="İşlenen ürün/parça adı zorunludur")
    if body.operation_type not in {"Tornalama", "Frezeleme", "Delme", "Diş açma", "Diğer"}:
        raise HTTPException(status_code=400, detail="Geçersiz iş türü")
    doc = {**body.model_dump(), "id": new_id(), "created_at": now_utc().isoformat(), "created_by": user.get("name", "")}
    if body.machine_id:
        machine = await db.machines.find_one({"id": body.machine_id})
        if not machine:
            raise HTTPException(status_code=404, detail="Tezgah bulunamadı")
        doc["machine_name"] = machine.get("name", "")
    await db.tool_trials.insert_one(doc)
    return clean(doc)


@api.delete("/tool-trials/{tid}")
async def delete_tool_trial(tid: str, user=Depends(require_admin)):
    r = await db.tool_trials.delete_one({"id": tid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Deneme kaydı bulunamadı")
    return {"ok": True}


@api.get("/tool-trials/compare")
async def compare_tool_trials(user=Depends(get_current_user), comparison_name: str = ""):
    q = {"comparison_name": comparison_name} if comparison_name else {}
    rows = await db.tool_trials.find(q, {"_id": 0}).sort("parts_machined", -1).to_list(5000)
    return {"comparison_name": comparison_name, "count": len(rows), "rows": rows}


# ---------- Machining recipes / İşleme reçeteleri ----------
@api.get("/recipes")
async def list_recipes(user=Depends(get_current_user)):
    return await db.recipes.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.post("/recipes")
async def create_recipe(body: RecipeIn, user=Depends(require_admin)):
    if not body.name.strip() or not body.workpiece_name.strip():
        raise HTTPException(status_code=400, detail="Reçete adı ve işlenecek ürün zorunludur")
    if not body.lines:
        raise HTTPException(status_code=400, detail="Reçeteye en az bir takım veya tutucu ekleyin")
    lines = []
    for line in body.lines:
        if line.kind not in {"product", "toolholder"} or line.quantity <= 0:
            raise HTTPException(status_code=400, detail="Geçersiz reçete satırı")
        collection = db.products if line.kind == "product" else db.toolholders
        item = await collection.find_one({"id": line.reference_id})
        if not item:
            raise HTTPException(status_code=404, detail="Reçete satırındaki stok kaydı bulunamadı")
        lines.append({"kind": line.kind, "reference_id": item["id"], "name": item.get("name", ""), "code": item.get("code", ""), "quantity": line.quantity, "note": line.note or ""})
    doc = {"id": new_id(), "name": body.name.strip(), "workpiece_name": body.workpiece_name.strip(), "note": body.note or "", "lines": lines, "status": "draft", "created_at": now_utc().isoformat(), "created_by": user.get("name", "")}
    await db.recipes.insert_one(doc)
    return clean(doc)


@api.post("/recipes/{rid}/bind")
async def bind_recipe(rid: str, body: RecipeBindIn, user=Depends(require_admin)):
    recipe = await db.recipes.find_one({"id": rid})
    if not recipe:
        raise HTTPException(status_code=404, detail="Reçete bulunamadı")
    if recipe.get("status") == "bound":
        raise HTTPException(status_code=400, detail="Bu reçete daha önce tezgâha bağlanmış")
    machine = await db.machines.find_one({"id": body.machine_id})
    if not machine:
        raise HTTPException(status_code=404, detail="Tezgah bulunamadı")
    resolved = []
    for line in recipe.get("lines", []):
        collection = db.products if line.get("kind") == "product" else db.toolholders
        item = await collection.find_one({"id": line.get("reference_id")})
        if not item:
            raise HTTPException(status_code=404, detail=f"Stok kaydı bulunamadı: {line.get('name', '')}")
        if line.get("quantity", 0) > item.get("current_stock", 0):
            raise HTTPException(status_code=400, detail=f"Yetersiz stok: {line.get('name', '')}")
        resolved.append((line, collection, item))
    movements = []
    for line, collection, item in resolved:
        new_stock = item.get("current_stock", 0) - line["quantity"]
        await collection.update_one({"id": item["id"]}, {"$set": {"current_stock": new_stock}})
        movement = {"id": new_id(), "type": "recipe_bind", "recipe_id": rid, "recipe_name": recipe["name"], "kind": line["kind"], "reference_id": item["id"], "product_code": item.get("code", ""), "product_name": item.get("name", ""), "quantity": line["quantity"], "machine_id": machine["id"], "machine_name": machine.get("name", ""), "user_id": user["id"], "user_name": user.get("name", ""), "created_at": now_utc().isoformat()}
        await db.recipe_movements.insert_one(movement)
        movements.append(clean(movement))
    await db.recipes.update_one({"id": rid}, {"$set": {"status": "bound", "machine_id": machine["id"], "machine_name": machine.get("name", ""), "bound_at": now_utc().isoformat(), "bound_by": user.get("name", "")}})
    updated = await db.recipes.find_one({"id": rid}, {"_id": 0})
    return {"ok": True, "recipe": updated, "movements": movements}


@api.get("/recipe-movements")
async def list_recipe_movements(user=Depends(get_current_user), recipe_id: Optional[str] = None):
    q = {"recipe_id": recipe_id} if recipe_id else {}
    return await db.recipe_movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)


# ---------- Dashboard ----------
@api.get("/dashboard")
async def dashboard(user=Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(2000)
    total_products = len(products)
    critical = [p for p in products if p.get("current_stock", 0) <= p.get("min_stock", 0)]
    critical_count = len(critical)

    month_start = now_utc().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_movements = await db.movements.find(
        {"type": "out", "created_at": {"$gte": month_start.isoformat()}}, {"_id": 0}
    ).to_list(5000)
    month_total = sum(m.get("total", 0) for m in month_movements)

    p_totals, m_totals = {}, {}
    for mv in month_movements:
        pn = mv.get("personnel_name", "-")
        mn = mv.get("machine_name", "-")
        p_totals[pn] = p_totals.get(pn, 0) + mv.get("quantity", 0)
        m_totals[mn] = m_totals.get(mn, 0) + mv.get("quantity", 0)
    top_personnel = sorted(p_totals.items(), key=lambda x: -x[1])[:5]
    top_machines = sorted(m_totals.items(), key=lambda x: -x[1])[:5]

    recent = await db.movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(10)
    personnel_list = await db.personnel.find({}, {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "department": 1}).sort("first_name", 1).to_list(2000)
    machine_list = await db.machines.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1, "brand": 1, "model": 1}).sort("code", 1).to_list(2000)
    personnel_totals = {name: round(qty, 2) for name, qty in p_totals.items()}
    machine_totals = {name: round(qty, 2) for name, qty in m_totals.items()}
    for person in personnel_list:
        person["name"] = f"{person.get('first_name', '')} {person.get('last_name', '')}".strip()
        person["qty"] = personnel_totals.get(person["name"], 0)
    for machine in machine_list:
        machine["qty"] = machine_totals.get(machine.get("name", ""), 0)

    return {
        "total_products": total_products,
        "critical_count": critical_count,
        "month_total_cost": 0,
        "critical_products": critical[:10],
        "top_personnel": [{"name": n, "qty": round(q, 2)} for n, q in top_personnel],
        "top_machines": [{"name": n, "qty": round(q, 2)} for n, q in top_machines],
        "personnel_list": personnel_list,
        "machine_list": machine_list,
        "recent_movements": recent,
    }


# ---------- Consumption detail reports ----------
@api.get("/reports/consumption-detail")
async def consumption_detail(user=Depends(get_current_user), personnel_id: Optional[str] = None, machine_id: Optional[str] = None):
    if not personnel_id and not machine_id:
        raise HTTPException(status_code=400, detail="Personel veya tezgâh seçilmelidir")
    movement_filter = {"type": "out"}
    if personnel_id:
        movement_filter["personnel_id"] = personnel_id
    if machine_id:
        movement_filter["machine_id"] = machine_id
    rows = await db.movements.find(movement_filter, {"_id": 0}).sort("created_at", -1).to_list(50000)
    machine_ids = {row.get("machine_id") for row in rows if row.get("machine_id")}
    machine_docs = await db.machines.find({"id": {"$in": list(machine_ids)}}, {"_id": 0, "id": 1, "code": 1, "name": 1}).to_list(2000) if machine_ids else []
    machine_map = {m.get("id"): m for m in machine_docs}
    for row in rows:
        machine = machine_map.get(row.get("machine_id"), {})
        row["machine_code"] = row.get("machine_code") or machine.get("code", "")
        row["machine_name"] = row.get("machine_name") or machine.get("name", "-")
    totals = {}
    for row in rows:
        transaction_date = row.get("transaction_date") or (row.get("created_at", "")[:10] if row.get("created_at") else "-")
        key = (transaction_date, row.get("product_id", ""), row.get("product_code", ""), row.get("product_name", "-"), row.get("machine_id", ""), row.get("machine_code", ""), row.get("machine_name", "-"), row.get("production_product", "") or "-")
        item = totals.setdefault(key, {"transaction_date": key[0], "personnel_name": row.get("personnel_name", "-"), "product_id": key[1], "product_code": key[2], "product_name": key[3], "machine_id": key[4], "machine_code": key[5], "machine_name": key[6], "production_product": key[7], "quantity": 0, "movement_count": 0})
        item["quantity"] += float(row.get("quantity", 0) or 0)
        item["movement_count"] += 1
    return {"rows": list(totals.values()), "movements": rows}


# ---------- Reports (Excel) ----------
def _date_query(date_from, date_to):
    q = {}
    if date_from or date_to:
        transaction_query = {}
        if date_from:
            transaction_query["$gte"] = date_from
        if date_to:
            transaction_query["$lte"] = date_to
        q["$or"] = [
            {"transaction_date": transaction_query},
            {"transaction_date": {"$exists": False}, "created_at": {"$gte": date_from or "", "$lte": (date_to + "T23:59:59") if date_to else "9999-12-31T23:59:59"}},
        ]
    return q


@api.get("/reports/summary")
async def report_summary(user=Depends(get_current_user),
                        date_from: Optional[str] = None, date_to: Optional[str] = None):
    q = _date_query(date_from, date_to)
    movements = await db.movements.find(q, {"_id": 0}).to_list(20000)

    by_product, by_personnel, by_machine = {}, {}, {}
    detail_map = {}
    for m in movements:
        if m.get("type") != "out":
            continue
        pk = m.get("product_name", "-")
        by_product.setdefault(pk, {"qty": 0, "total": 0, "code": m.get("product_code", "")})
        by_product[pk]["qty"] += m.get("quantity", 0)
        by_product[pk]["total"] += m.get("total", 0)
        per = m.get("personnel_name", "-")
        by_personnel.setdefault(per, {"qty": 0, "total": 0})
        by_personnel[per]["qty"] += m.get("quantity", 0)
        by_personnel[per]["total"] += m.get("total", 0)
        mc = m.get("machine_name", "-")
        by_machine.setdefault(mc, {"qty": 0, "total": 0})
        by_machine[mc]["qty"] += m.get("quantity", 0)
        by_machine[mc]["total"] += m.get("total", 0)
        detail_key = (per, mc, m.get("product_name", "-"), m.get("production_product", "-") or "-")
        detail_map.setdefault(detail_key, {"personnel": per, "machine": mc, "tool": m.get("product_name", "-"), "code": m.get("product_code", ""), "production_product": m.get("production_product", "-") or "-", "qty": 0})
        detail_map[detail_key]["qty"] += m.get("quantity", 0)

    return {
        "by_product": [{"name": k, **v} for k, v in by_product.items()],
        "by_personnel": [{"name": k, **v} for k, v in by_personnel.items()],
        "by_machine": [{"name": k, **v} for k, v in by_machine.items()],
        "details": list(detail_map.values()),
        "count": len(movements),
    }


@api.get("/reports/excel")
async def report_excel(user=Depends(get_current_user),
                       date_from: Optional[str] = None, date_to: Optional[str] = None):
    q = _date_query(date_from, date_to)
    movements = await db.movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(50000)
    # Bilemeye gönderilen kayıtlar ayrı koleksiyonda tutulduğu için raporun
    # Hareketler sayfasına açıklayıcı bir hareket satırı olarak eklenir.
    sharpening_records = await db.sharpening_records.find(q, {"_id": 0}).sort("created_at", -1).to_list(50000)
    def _same_sharpening_movement(sr, m):
        if m.get("type") != "out":
            return False
        if sr.get("id") and m.get("sharpening_record_id") == sr.get("id"):
            return True
        if m.get("product_id") != sr.get("product_id"):
            return False
        try:
            if float(m.get("quantity", 0)) != float(sr.get("quantity", 0)):
                return False
        except (TypeError, ValueError):
            return False
        sent_day = str(sr.get("sent_date", ""))[:10]
        movement_day = str(m.get("transaction_date") or m.get("created_at", ""))[:10]
        if not sent_day or sent_day != movement_day:
            return False
        # Eski sürümlerde bileme önce normal stok çıkışı, sonra bileme kaydı olarak
        # tutulmuş olabilir. Önce kesin kayıt bağı ve işlem zamanı kontrol edilir;
        # eski kayıtlarda zaman bağı yoksa aynı gün/miktar ve boş üretim alanı
        # bileme için oluşturulmuş stok çıkışını tekilleştirmek için kullanılır.
        try:
            created_a = datetime.fromisoformat(str(sr.get("created_at", "")).replace("Z", "+00:00"))
            created_b = datetime.fromisoformat(str(m.get("created_at", "")).replace("Z", "+00:00"))
            if created_a.tzinfo is None:
                created_a = created_a.replace(tzinfo=timezone.utc)
            if created_b.tzinfo is None:
                created_b = created_b.replace(tzinfo=timezone.utc)
            if abs((created_a - created_b).total_seconds()) <= 15 * 60:
                return True
        except (TypeError, ValueError, OverflowError):
            pass
        note = f"{m.get('note', '')} {m.get('production_product', '')}".casefold()
        return "bileme" in note or (not m.get("production_product") and not m.get("note"))

    for sr in sharpening_records:
        if any(_same_sharpening_movement(sr, m) for m in movements):
            continue
        movements.append({
            "created_at": sr.get("created_at", ""),
            "transaction_date": sr.get("sent_date", ""),
            "type": "out",
            "product_code": sr.get("product_code", ""),
            "product_name": sr.get("product_name", ""),
            "quantity": sr.get("quantity", 0),
            "personnel_name": sr.get("sent_by", ""),
            "machine_name": "",
            "production_product": "",
            "supplier": sr.get("company", ""),
            "note": sr.get("note", ""),
            "user_name": sr.get("sent_by", ""),
            "sharpening_record_id": sr.get("id", ""),
            "movement_purpose": "Bilemeye gitti",
            "destination": sr.get("company", ""),
            "process_type": sr.get("process_type", ""),
            "helix_length": sr.get("helix_length", ""),
            "diameter": sr.get("diameter", ""),
            "full_length": sr.get("full_length", ""),
        })

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Hareketler"
    ws1.append(["Tarih", "Tip", "İşlem / Amaç", "Çıkış / Hedef", "Bileme İşlemi", "Ürün Kodu", "Ürün Adı", "Miktar",
                "Personel", "Tezgah", "Üretim / İşlenen Ürün", "Helis", "Çap", "Tam Boy", "Tedarikçi", "Not", "Kullanıcı"])
    for m in movements:
        ws1.append([
            m.get("transaction_date") or m.get("created_at", "")[:19].replace("T", " "),
            "GİRİŞ" if m.get("type") == "in" else "ÇIKIŞ",
            m.get("movement_purpose") or ("Bilemeye gitti" if m.get("sharpening_record_id") else ("Stok girişi" if m.get("type") == "in" else ("İşleme için verildi" if m.get("machine_name") else "Üretimde kullanım"))),
            m.get("destination") or (m.get("machine_name", "") if m.get("type") == "out" else m.get("supplier", "")),
            m.get("process_type", ""), m.get("product_code", ""), m.get("product_name", ""),
            m.get("quantity", 0), m.get("personnel_name", ""), m.get("machine_name", ""),
            m.get("production_product", ""), m.get("helix_length", ""), m.get("diameter", ""), m.get("full_length", ""),
            m.get("supplier", ""), m.get("note", ""), m.get("user_name", ""),
        ])
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    def _summary(sheet, key, header):
        ws = wb.create_sheet(sheet)
        ws.append(header)
        agg = {}
        for m in movements:
            if m.get("type") != "out":
                continue
            k = m.get(key, "-")
            agg[k] = agg.get(k, 0) + m.get("quantity", 0)
        for k, v in agg.items():
            ws.append([k, v])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    _summary("Ürün Bazlı", "product_name", ["Ürün", "Toplam Miktar"])

    detail_headers = ["İşlem Tarihi", "Personel", "Tezgah", "Kullanılan Uç / Ürün", "Kod", "Üretim / İşlenen Ürün", "Miktar", "Amaç"]
    def _detail_sheet(sheet, sort_key):
        ws = wb.create_sheet(sheet)
        ws.append(detail_headers)
        rows = [m for m in movements if m.get("type") == "out"]
        rows.sort(key=lambda m: (m.get(sort_key, ""), m.get("transaction_date") or m.get("created_at", "")))
        for m in rows:
            ws.append([
                m.get("transaction_date") or m.get("created_at", "")[:10],
                m.get("personnel_name", ""), m.get("machine_name", ""), m.get("product_name", ""),
                m.get("product_code", ""), m.get("production_product", ""), m.get("quantity", 0),
                "Bilemeye gönderildi" if m.get("sharpening_record_id") else "Üretimde kullanım",
            ])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    _detail_sheet("Personel Bazlı", "personnel_name")
    _detail_sheet("Tezgah Bazlı", "machine_name")

    detail_ws = wb.create_sheet("Kullanım Detayı")
    detail_ws.append(["Tarih", "Personel", "Tezgah", "Kullanılan Uç / Ürün", "Kod", "Üretim / İşlenen Ürün", "Miktar", "Amaç"])
    for m in movements:
        if m.get("type") != "out":
            continue
        detail_ws.append([
            m.get("transaction_date") or m.get("created_at", "")[:10],
            m.get("personnel_name", ""), m.get("machine_name", ""), m.get("product_name", ""),
            m.get("product_code", ""), m.get("production_product", ""), m.get("quantity", 0),
            "Bilemeye gönderildi" if m.get("sharpening_record_id") else "Üretimde kullanım",
        ])
    detail_ws.freeze_panes = "A2"
    detail_ws.auto_filter.ref = detail_ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"rapor_{now_utc().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origin_regex=".*",
    allow_methods=["*"],
    allow_headers=["*"],
)


# ==================== Suppliers ====================
@api.get("/suppliers")
async def list_suppliers(user=Depends(get_current_user)):
    return await db.suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)


@api.post("/suppliers")
async def create_supplier(body: SupplierIn, user=Depends(require_admin)):
    if await db.suppliers.find_one({"name": body.name}):
        raise HTTPException(status_code=400, detail="Bu isimde tedarikçi zaten mevcut")
    doc = {**body.model_dump(), "id": new_id(), "created_at": now_utc().isoformat()}
    await db.suppliers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/suppliers/{sid}")
async def update_supplier(sid: str, body: SupplierIn, user=Depends(require_admin)):
    existing = await db.suppliers.find_one({"id": sid})
    if not existing:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")
    if body.name != existing["name"] and await db.suppliers.find_one({"name": body.name}):
        raise HTTPException(status_code=400, detail="Bu isimde tedarikçi zaten mevcut")
    await db.suppliers.update_one({"id": sid}, {"$set": body.model_dump()})
    return await db.suppliers.find_one({"id": sid}, {"_id": 0})


@api.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, user=Depends(require_admin)):
    r = await db.suppliers.delete_one({"id": sid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")
    return {"ok": True}


@api.get("/reports/by-supplier")
async def report_by_supplier(user=Depends(get_current_user),
                             date_from: Optional[str] = None, date_to: Optional[str] = None):
    q = _date_query(date_from, date_to)
    q["type"] = "in"
    movements = await db.movements.find(q, {"_id": 0}).to_list(20000)
    agg = {}
    for m in movements:
        key = m.get("supplier") or "Belirtilmemiş"
        agg.setdefault(key, {"qty": 0, "total": 0, "count": 0})
        agg[key]["qty"] += m.get("quantity", 0)
        agg[key]["total"] += m.get("total", 0)
        agg[key]["count"] += 1
    return [{"name": k, **v} for k, v in sorted(agg.items(), key=lambda x: -x[1]["total"])]


# ==================== Tool Holders (Takım Tutucular) ====================
@api.get("/toolholders")
async def list_toolholders(user=Depends(get_current_user)):
    return await db.toolholders.find({}, {"_id": 0}).sort("name", 1).to_list(2000)


@api.post("/toolholders")
async def create_toolholder(body: ToolHolderIn_Model, user=Depends(require_admin)):
    doc = {**body.model_dump(), "id": new_id(), "created_at": now_utc().isoformat()}
    await db.toolholders.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/toolholders/{tid}")
async def update_toolholder(tid: str, body: ToolHolderIn_Model, user=Depends(require_admin)):
    existing = await db.toolholders.find_one({"id": tid})
    if not existing:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    await db.toolholders.update_one({"id": tid}, {"$set": body.model_dump()})
    return await db.toolholders.find_one({"id": tid}, {"_id": 0})


@api.delete("/toolholders/{tid}")
async def delete_toolholder(tid: str, user=Depends(require_admin)):
    r = await db.toolholders.delete_one({"id": tid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    return {"ok": True}


@api.post("/toolholders/{tid}/in")
async def toolholder_in(tid: str, body: ToolHolderStockIn, user=Depends(require_admin)):
    """Takımhaneye giriş: stok artar."""
    h = await db.toolholders.find_one({"id": tid})
    if not h:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    new_stock = h.get("current_stock", 0) + body.quantity
    await db.toolholders.update_one({"id": tid}, {"$set": {"current_stock": new_stock}})
    mv = {
        "id": new_id(), "type": "in", "tool_holder_id": h["id"], "name": h["name"],
        "brand": h.get("brand", ""), "holder_type": h.get("type", ""),
        "length": h.get("length", ""), "diameter": h.get("diameter", ""),
        "quantity": body.quantity,
        "supplier": body.supplier or "", "supplier_id": body.supplier_id or "",
        "note": body.note or "",
        "user_id": user["id"], "user_name": user["name"],
        "created_at": now_utc().isoformat(),
    }
    await db.toolholder_movements.insert_one(mv)
    return {"ok": True, "new_stock": new_stock, "movement": clean(mv)}


@api.post("/toolholders/{tid}/out")
async def toolholder_out(tid: str, body: ToolHolderStockOut, user=Depends(require_admin)):
    """Tezgaha çıkış: hangi tezgaha bağlandığı bilgisi ile stoktan düşer."""
    h = await db.toolholders.find_one({"id": tid})
    if not h:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Miktar 0'dan büyük olmalı")
    if body.quantity > h.get("current_stock", 0):
        raise HTTPException(status_code=400, detail="Yetersiz stok")
    machine = await db.machines.find_one({"id": body.machine_id})
    if not machine:
        raise HTTPException(status_code=404, detail="Tezgah bulunamadı")
    personnel = None
    if body.personnel_id:
        personnel = await db.personnel.find_one({"id": body.personnel_id})
    new_stock = h["current_stock"] - body.quantity
    await db.toolholders.update_one({"id": tid}, {"$set": {"current_stock": new_stock}})
    mv = {
        "id": new_id(), "type": "out", "tool_holder_id": h["id"], "name": h["name"],
        "brand": h.get("brand", ""), "holder_type": h.get("type", ""),
        "length": h.get("length", ""), "diameter": h.get("diameter", ""),
        "quantity": body.quantity,
        "machine_id": machine["id"], "machine_name": machine["name"],
        "machine_code": machine.get("code", ""),
        "personnel_id": personnel["id"] if personnel else "",
        "personnel_name": f"{personnel['first_name']} {personnel['last_name']}" if personnel else "",
        "note": body.note or "",
        "user_id": user["id"], "user_name": user["name"],
        "created_at": now_utc().isoformat(),
    }
    await db.toolholder_movements.insert_one(mv)
    return {"ok": True, "new_stock": new_stock, "movement": clean(mv)}


@api.get("/toolholder-scraps")
async def list_toolholder_scraps(user=Depends(get_current_user), limit: int = 500):
    return await db.toolholder_scraps.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)


@api.post("/toolholders/{tid}/scrap")
async def toolholder_scrap(tid: str, body: ToolHolderScrapIn, user=Depends(require_admin)):
    holder = await db.toolholders.find_one({"id": tid})
    if not holder:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Hurda miktarı 0'dan büyük olmalı")
    if body.quantity > float(holder.get("current_stock", 0)):
        raise HTTPException(status_code=400, detail="Hurda miktarı mevcut stoktan fazla olamaz")
    if not body.scrap_reason.strip():
        raise HTTPException(status_code=400, detail="Hurda nedeni zorunludur")
    new_stock = round(float(holder.get("current_stock", 0)) - body.quantity, 6)
    async def personnel_name(pid: str) -> str:
        if not pid:
            return ""
        person = await db.personnel.find_one({"id": pid})
        return f"{person.get('first_name', '')} {person.get('last_name', '')}".strip() if person else ""
    operator_name = await personnel_name(body.operator_personnel_id or "") or user.get("name", "")
    approver_name = await personnel_name(body.approver_personnel_id or "") or body.approved_by or ""
    witness_name = await personnel_name(body.witness_personnel_id or "") or body.witness or ""
    now = now_utc().isoformat()
    doc = {
        "id": new_id(), "tool_holder_id": holder["id"], "name": holder.get("name", ""),
        "brand": holder.get("brand", ""), "holder_type": holder.get("type", ""),
        "cutting_tool_code_name": holder.get("cutting_tool_code_name", ""),
        "length": holder.get("length", ""), "diameter": holder.get("diameter", ""),
        "quantity": body.quantity, "unit": "adet", "scrap_reason": body.scrap_reason.strip(),
        "description": body.description or "", "scrap_date": body.scrap_date,
        "location": body.location or holder.get("location", ""), "approved_by": approver_name,
        "witness": witness_name, "operator_personnel_id": body.operator_personnel_id or "",
        "approver_personnel_id": body.approver_personnel_id or "", "witness_personnel_id": body.witness_personnel_id or "",
        "operator_name": operator_name, "user_id": user["id"], "user_name": user.get("name", ""),
        "created_at": now, "new_stock": new_stock,
    }
    await db.toolholders.update_one({"id": tid}, {"$set": {"current_stock": new_stock}})
    await db.toolholder_scraps.insert_one(doc)
    await db.toolholder_movements.insert_one({
        "id": new_id(), "type": "scrap", "tool_holder_id": holder["id"], "name": holder.get("name", ""),
        "quantity": body.quantity, "scrap_reason": body.scrap_reason.strip(), "note": body.description or "",
        "user_id": user["id"], "user_name": user.get("name", ""), "created_at": now,
    })
    return {"ok": True, "new_stock": new_stock, "scrap": clean(doc)}


@api.put("/toolholder-scraps/{scrap_id}")
async def update_toolholder_scrap(scrap_id: str, body: ToolHolderScrapIn, user=Depends(require_admin)):
    record = await db.toolholder_scraps.find_one({"id": scrap_id})
    if not record:
        raise HTTPException(status_code=404, detail="Hurda kaydı bulunamadı")
    if body.quantity <= 0 or not body.scrap_reason.strip():
        raise HTTPException(status_code=400, detail="Miktar ve hurda nedeni zorunludur")
    holder = await db.toolholders.find_one({"id": record["tool_holder_id"]})
    if not holder:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    old_qty = float(record.get("quantity", 0))
    current_stock = float(holder.get("current_stock", 0))
    new_stock = round(current_stock + old_qty - body.quantity, 6)
    if new_stock < 0:
        raise HTTPException(status_code=400, detail="Bu düzeltme stok miktarını eksiye düşürüyor")
    async def personnel_name(pid: str) -> str:
        if not pid:
            return ""
        person = await db.personnel.find_one({"id": pid})
        return f"{person.get('first_name', '')} {person.get('last_name', '')}".strip() if person else ""
    operator_name = await personnel_name(body.operator_personnel_id or "") or record.get("operator_name", "") or user.get("name", "")
    approver_name = await personnel_name(body.approver_personnel_id or "") or body.approved_by or ""
    witness_name = await personnel_name(body.witness_personnel_id or "") or body.witness or ""
    update = {
        "quantity": body.quantity, "scrap_reason": body.scrap_reason.strip(), "description": body.description or "",
        "scrap_date": body.scrap_date, "location": body.location or holder.get("location", ""),
        "operator_personnel_id": body.operator_personnel_id or "", "approver_personnel_id": body.approver_personnel_id or "",
        "witness_personnel_id": body.witness_personnel_id or "", "operator_name": operator_name,
        "approved_by": approver_name, "witness": witness_name, "new_stock": new_stock,
        "updated_at": now_utc().isoformat(), "updated_by": user.get("name", ""),
    }
    await db.toolholders.update_one({"id": holder["id"]}, {"$set": {"current_stock": new_stock}})
    await db.toolholder_scraps.update_one({"id": scrap_id}, {"$set": update})
    await db.toolholder_movements.insert_one({
        "id": new_id(), "type": "scrap_adjustment", "tool_holder_id": holder["id"], "name": holder.get("name", ""),
        "quantity": round(body.quantity - old_qty, 6), "note": "Hurda kaydı düzeltildi", "scrap_id": scrap_id,
        "user_id": user["id"], "user_name": user.get("name", ""), "created_at": now_utc().isoformat(),
    })
    updated = await db.toolholder_scraps.find_one({"id": scrap_id}, {"_id": 0})
    return {"ok": True, "new_stock": new_stock, "scrap": updated}


@api.delete("/toolholder-scraps/{scrap_id}")
async def delete_toolholder_scrap(scrap_id: str, user=Depends(require_admin)):
    record = await db.toolholder_scraps.find_one({"id": scrap_id})
    if not record:
        raise HTTPException(status_code=404, detail="Hurda kaydı bulunamadı")
    holder = await db.toolholders.find_one({"id": record["tool_holder_id"]})
    if not holder:
        raise HTTPException(status_code=404, detail="Tutucu bulunamadı")
    new_stock = round(float(holder.get("current_stock", 0)) + float(record.get("quantity", 0)), 6)
    await db.toolholders.update_one({"id": holder["id"]}, {"$set": {"current_stock": new_stock}})
    await db.toolholder_scraps.delete_one({"id": scrap_id})
    await db.toolholder_movements.insert_one({
        "id": new_id(), "type": "scrap_cancel", "tool_holder_id": holder["id"], "name": holder.get("name", ""),
        "quantity": record.get("quantity", 0), "note": "Hurda kaydı silindi; stok geri eklendi", "scrap_id": scrap_id,
        "user_id": user["id"], "user_name": user.get("name", ""), "created_at": now_utc().isoformat(),
    })
    return {"ok": True, "new_stock": new_stock}


@api.get("/toolholder-scraps/{scrap_id}/pdf")
async def toolholder_scrap_pdf(scrap_id: str, user=Depends(get_current_user)):
    scrap = await db.toolholder_scraps.find_one({"id": scrap_id}, {"_id": 0})
    if not scrap:
        raise HTTPException(status_code=404, detail="Hurda kaydı bulunamadı")
    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("DejaVu", font_path))
        pdf.setFont("DejaVu", 11)
    else:
        pdf.setFont("Helvetica", 11)
    width, height = A4
    y = height - 60
    pdf.setFont("DejaVu" if os.path.exists(font_path) else "Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, y, "TAKIM TUTUCU HURDA / KULLANIM DIŞI TUTANAĞI")
    y -= 35
    pdf.setFont("DejaVu" if os.path.exists(font_path) else "Helvetica", 10)
    rows = [
        ("Tutanak No", scrap.get("id", "")), ("Tarih", scrap.get("scrap_date", "")),
        ("Tutucu", scrap.get("name", "")), ("Marka", scrap.get("brand", "")),
        ("Tutucu Tipi", scrap.get("holder_type", "")), ("Kesici Uç Kodu / İsmi", scrap.get("cutting_tool_code_name", "")),
        ("Ölçüler", f"Boy: {scrap.get('length', '')} | Çap: {scrap.get('diameter', '')}"),
        ("Hurda Miktarı", f"{scrap.get('quantity', 0)} {scrap.get('unit', 'adet')}"),
        ("Hurda Nedeni", scrap.get("scrap_reason", "")), ("Açıklama", scrap.get("description", "")),
        ("Konum", scrap.get("location", "")), ("İşlemi Yapan", scrap.get("user_name", "")),
        ("Onaylayan", scrap.get("approved_by", "")), ("Tanık / Teslim Alan", scrap.get("witness", "")),
    ]
    for label, value in rows:
        pdf.setFont("DejaVu" if os.path.exists(font_path) else "Helvetica-Bold", 10)
        pdf.drawString(55, y, f"{label}:")
        pdf.setFont("DejaVu" if os.path.exists(font_path) else "Helvetica", 10)
        pdf.drawString(215, y, str(value)[:100])
        y -= 22
        if y < 100:
            pdf.showPage(); y = height - 60
    image_rows = [("Eski Hâli / Stoktaki Görsel", scrap.get("before_image_url", "")), ("Yeni Hasarlı Hâli", scrap.get("after_image_url", ""))]
    for caption, data_url in image_rows:
        if not data_url or not str(data_url).startswith("data:image/"):
            continue
        try:
            raw = base64.b64decode(str(data_url).split(",", 1)[1])
            img = ImageReader(io.BytesIO(raw))
            if y < 240:
                pdf.showPage(); y = height - 60
            pdf.setFont("DejaVu" if os.path.exists(font_path) else "Helvetica-Bold", 10)
            pdf.drawString(55, y, caption)
            y -= 10
            pdf.drawImage(img, 55, y - 175, width=220, height=165, preserveAspectRatio=True, anchor="sw", mask="auto")
            y -= 195
        except Exception:
            continue
    y -= 15
    pdf.drawString(70, y, "Düzenleyen İmza: ____________________")
    pdf.drawString(330, y, "Onay İmza: ____________________")
    pdf.save(); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="hurda-tutanagi-{scrap_id[:8]}.pdf"'})


@api.get("/toolholder-movements")
async def list_toolholder_movements(user=Depends(get_current_user),
                                    tool_holder_id: Optional[str] = None,
                                    type: Optional[str] = None,
                                    limit: int = 500):
    q = {}
    if tool_holder_id:
        q["tool_holder_id"] = tool_holder_id
    if type:
        q["type"] = type
    return await db.toolholder_movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)


# ==================== Tool Holder bulk import ====================
@api.get("/toolholders/import/template")
async def toolholder_import_template(user=Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tutucular"
    ws.append(["name", "brand", "cutting_tool_code_name", "type", "length", "diameter", "min_stock", "current_stock", "location", "note", "image_url"])
    ws.append(["BT40 Örnek Tutucu", "Regofix", "WNMG 080412-PM", "BT40", "120", "32", 1, 5, "Raf A-1", "", ""])
    ws.append(["HSK-A63 ER32", "Big Kaiser", "ER32 pens seti", "HSK-A63", "90", "25", 1, 3, "Raf A-2", "Kısa tip", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tutucu_sablonu.xlsx"'},
    )


@api.post("/toolholders/import")
async def toolholder_import(file: UploadFile = File(...), commit: bool = False,
                            user=Depends(require_admin)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyası yükleyin")
    content = await file.read()
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Dosyada veri yok")
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h}
    if "name" not in idx:
        raise HTTPException(status_code=400, detail="Zorunlu sütun eksik: name")

    def _get(row, col, default=None):
        i = idx.get(col)
        if i is None or i >= len(row):
            return default
        return row[i]

    def _safe_str(v):
        # datetime/int/float/None → clean string. Strips BOM whitespace.
        if v is None:
            return ""
        return str(v).strip().lstrip("\ufeff")

    def _safe_num(v, default=0.0):
        # NaN / Inf / bad text → default. Otherwise finite float.
        if v is None or (isinstance(v, str) and not v.strip()):
            return default
        try:
            # Turkish decimal comma tolerated
            if isinstance(v, str):
                v = v.replace(",", ".").strip()
            x = float(v)
        except (TypeError, ValueError):
            raise
        if x != x:  # NaN
            return default
        if x == float("inf") or x == float("-inf"):
            return default
        return x

    preview = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or _safe_str(v) == "" for v in row):
            continue
        name = _safe_str(_get(row, "name"))
        if not name:
            preview.append({"row": row_num, "action": "skip", "error": "name boş olamaz",
                            "data": {"name": name}})
            continue
        try:
            min_stock = _safe_num(_get(row, "min_stock"), 0.0)
            current_stock = _safe_num(_get(row, "current_stock"), 0.0)
        except (TypeError, ValueError):
            preview.append({"row": row_num, "action": "skip",
                            "error": "min_stock/current_stock sayı olmalı",
                            "data": {"name": name}})
            continue
        d = {
            "name": name,
            "brand": _safe_str(_get(row, "brand")),
            "cutting_tool_code_name": _safe_str(_get(row, "cutting_tool_code_name")),
            "type": _safe_str(_get(row, "type")),
            "length": _safe_str(_get(row, "length")),
            "diameter": _safe_str(_get(row, "diameter")),
            "min_stock": min_stock,
            "current_stock": current_stock,
            "location": _safe_str(_get(row, "location")),
            "note": _safe_str(_get(row, "note")),
            "image_url": _safe_str(_get(row, "image_url")),
        }
        # Match on name + brand + type (best-effort uniqueness key)
        query = {"name": d["name"]}
        if d["brand"]:
            query["brand"] = d["brand"]
        if d["type"]:
            query["type"] = d["type"]
        existing = await db.toolholders.find_one(query)
        preview.append({"row": row_num,
                        "action": "update" if existing else "create",
                        "error": None, "data": d})

    stats = {
        "total": len(preview),
        "create": sum(1 for p in preview if p["action"] == "create"),
        "update": sum(1 for p in preview if p["action"] == "update"),
        "skip": sum(1 for p in preview if p["action"] == "skip"),
    }
    if not commit:
        return {"committed": False, "stats": stats, "preview": preview}

    created, updated = 0, 0
    for p in preview:
        if p["action"] == "skip":
            continue
        d = p["data"]
        query = {"name": d["name"]}
        if d["brand"]:
            query["brand"] = d["brand"]
        if d["type"]:
            query["type"] = d["type"]
        if p["action"] == "create":
            await db.toolholders.insert_one({**d, "id": new_id(),
                                             "created_at": now_utc().isoformat()})
            created += 1
        else:
            await db.toolholders.update_one(query, {"$set": d})
            updated += 1
    return {"committed": True, "stats": stats, "created": created,
            "updated": updated, "preview": preview}


# ==================== Admin: wipe all business data ====================
@api.post("/admin/wipe-all")
async def wipe_all_data(user=Depends(require_admin)):
    """Sadece admin: tüm işletme verilerini siler (ürünler, personel, tezgahlar, tedarikçiler, siparişler, hareketler, tutucular). Kullanıcı hesapları korunur. Ayrıca 'sample_seeded' işaretini set ederek yeniden başlatmada örnek veri tekrar eklenmesini önler."""
    collections = ["products", "personnel", "machines", "suppliers", "orders",
                   "movements", "toolholders", "toolholder_movements",
                   "password_resets"]
    result = {}
    for c in collections:
        r = await db[c].delete_many({})
        result[c] = r.deleted_count
    await db.settings.update_one({"key": "sample_seeded"},
                                 {"$set": {"key": "sample_seeded", "value": True,
                                           "at": now_utc().isoformat(),
                                           "reason": "wiped by admin"}},
                                 upsert=True)
    return {"ok": True, "deleted": result}


# ==================== Orders ====================
def _compute_order_totals(items):
    total = 0.0
    for it in items:
        total += (it.get("quantity", 0) or 0) * (it.get("unit_price", 0) or 0)
    return round(total, 2)


@api.get("/orders")
async def list_orders(user=Depends(get_current_user), status: Optional[str] = None):
    q = {}
    if status:
        q["status"] = status
    return await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.post("/orders")
async def create_order(body: OrderIn, user=Depends(require_admin)):
    supplier = await db.suppliers.find_one({"id": body.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")
    if not body.items:
        raise HTTPException(status_code=400, detail="En az bir kalem eklemelisiniz")
    items = []
    for it in body.items:
        if it.quantity <= 0:
            raise HTTPException(status_code=400, detail="Kalem miktarı 0'dan büyük olmalı")
        if it.product_id:
            prod = await db.products.find_one({"id": it.product_id})
            if not prod:
                raise HTTPException(status_code=404, detail=f"Ürün bulunamadı: {it.product_id}")
            items.append({
                "product_id": prod["id"], "product_code": prod["code"], "product_name": prod["name"],
                "category": prod.get("category", "Diğer"), "unit": prod.get("unit", "adet"),
                "quantity": it.quantity, "received_qty": 0, "manual": False,
            })
        else:
            name = (it.product_name or "").strip()
            if not name:
                raise HTTPException(status_code=400, detail="Manuel kalem için ürün adı gerekli")
            items.append({
                "product_id": None,
                "product_code": (it.product_code or "").strip(),
                "product_name": name,
                "category": (it.category or "Diğer").strip() or "Diğer",
                "unit": (it.unit or "adet").strip() or "adet",
                "quantity": it.quantity, "received_qty": 0, "manual": True,
            })
    order = {
        "id": new_id(),
        "supplier_id": supplier["id"], "supplier_name": supplier["name"],
        "delivery_date": body.delivery_date or "",
        "note": body.note or "",
        "status": "open",
        "items": items,
        "total": 0,
        "created_by": user["name"],
        "created_at": now_utc().isoformat(),
        "closed_at": None,
    }
    await db.orders.insert_one(order)
    order.pop("_id", None)
    return order


async def _ensure_product_for_item(item: dict) -> dict:
    """Return the product doc for an order item, creating it if manual + missing."""
    if item.get("product_id"):
        p = await db.products.find_one({"id": item["product_id"]})
        if p:
            return p
    # Try lookup by code
    code = (item.get("product_code") or "").strip()
    if code:
        p = await db.products.find_one({"code": code})
        if p:
            return p
    # Create a new product
    if not code:
        code = await _next_product_code()
    doc = {
        "id": new_id(),
        "code": code,
        "name": item.get("product_name") or code,
        "category": item.get("category") or "Diğer",
        "unit": item.get("unit") or "adet",
        "min_stock": 0.0, "current_stock": 0.0,
        "location": "", "quality": "", "brand": "",
        "created_at": now_utc().isoformat(),
    }
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.post("/orders/{oid}/close")
async def close_order(oid: str, user=Depends(require_admin)):
    order = await db.orders.find_one({"id": oid})
    if not order:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
    if order.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Sipariş zaten kapalı")
    # Convert each item to a stock-in movement and update product stock.
    # For manual items missing a product, auto-create one.
    new_items = [dict(it) for it in order.get("items", [])]
    for it in new_items:
        remaining = it["quantity"] - it.get("received_qty", 0)
        if remaining <= 0:
            continue
        prod = await _ensure_product_for_item(it)
        it["product_id"] = prod["id"]
        it["product_code"] = prod["code"]
        it["product_name"] = prod["name"]
        it["manual"] = False
        new_stock = prod.get("current_stock", 0) + remaining
        await db.products.update_one({"id": prod["id"]}, {"$set": {"current_stock": new_stock}})
        await db.movements.insert_one({
            "id": new_id(), "type": "in", "product_id": prod["id"],
            "product_code": prod["code"], "product_name": prod["name"],
            "quantity": remaining, "unit_price": 0, "total": 0,
            "supplier": order["supplier_name"], "supplier_id": order["supplier_id"],
            "note": f"Sipariş #{order['id'][:8]} kapatma",
            "order_id": order["id"],
            "user_id": user["id"], "user_name": user["name"],
            "created_at": now_utc().isoformat(),
        })
        it["received_qty"] = it["quantity"]
    await db.orders.update_one({"id": oid},
                               {"$set": {"status": "closed",
                                         "closed_at": now_utc().isoformat(),
                                         "items": new_items}})
    return await db.orders.find_one({"id": oid}, {"_id": 0})


@api.delete("/orders/{oid}")
async def delete_order(oid: str, user=Depends(require_admin)):
    r = await db.orders.delete_one({"id": oid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
    return {"ok": True}


@api.post("/orders/{oid}/receive")
async def receive_order(oid: str, body: ReceiveIn, user=Depends(require_admin)):
    order = await db.orders.find_one({"id": oid})
    if not order:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
    if order.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Sipariş zaten kapalı")
    if not body.items:
        raise HTTPException(status_code=400, detail="Teslim alınacak kalem yok")

    new_items = [dict(it) for it in order["items"]]
    any_received_now = False

    for rcv in body.items:
        if rcv.quantity <= 0:
            continue
        it = None
        if rcv.product_id:
            it = next((x for x in new_items if x.get("product_id") == rcv.product_id
                       or (x.get("product_id") is None and x.get("product_code") == rcv.product_id)), None)
        if it is None:
            # rcv.product_id may actually be the code for manual items
            it = next((x for x in new_items if x.get("product_code") == rcv.product_id), None)
        if not it:
            raise HTTPException(status_code=400, detail="Kalem sipariste yok")
        remaining = it["quantity"] - it.get("received_qty", 0)
        if rcv.quantity > remaining + 1e-9:
            raise HTTPException(status_code=400,
                                detail=f"{it['product_name']}: kalan {remaining} birim, fazla teslim alınamaz")
        it["received_qty"] = it.get("received_qty", 0) + rcv.quantity
        any_received_now = True

        prod = await _ensure_product_for_item(it)
        it["product_id"] = prod["id"]
        it["product_code"] = prod["code"]
        it["product_name"] = prod["name"]
        it["manual"] = False
        new_stock = prod.get("current_stock", 0) + rcv.quantity
        await db.products.update_one({"id": prod["id"]}, {"$set": {"current_stock": new_stock}})
        await db.movements.insert_one({
            "id": new_id(), "type": "in", "product_id": prod["id"],
            "product_code": prod["code"], "product_name": prod["name"],
            "quantity": rcv.quantity, "unit_price": 0, "total": 0,
            "supplier": order["supplier_name"], "supplier_id": order["supplier_id"],
            "note": f"Sipariş #{order['id'][:8]} kısmi teslimat",
            "order_id": order["id"],
            "user_id": user["id"], "user_name": user["name"],
            "created_at": now_utc().isoformat(),
        })

    if not any_received_now:
        raise HTTPException(status_code=400, detail="Geçerli teslim miktarı yok")

    all_received = all(it.get("received_qty", 0) >= it["quantity"] - 1e-9 for it in new_items)
    new_status = "closed" if all_received else "partial"
    update_doc = {"items": new_items, "status": new_status}
    if new_status == "closed":
        update_doc["closed_at"] = now_utc().isoformat()
    await db.orders.update_one({"id": oid}, {"$set": update_doc})
    return await db.orders.find_one({"id": oid}, {"_id": 0})


# ==================== Products bulk import ====================
@api.get("/products/import/template")
async def product_import_template(user=Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    ws.append(["code", "name", "category", "unit", "min_stock", "current_stock", "location", "quality", "brand", "image_url"])
    ws.append(["", "Örnek Kesici Uç", "Kesici Uç", "adet", 10, 25, "Raf A-1", "TiN", "Sandvik"])
    ws.append(["", "Örnek Matkap Ø8", "Matkap", "adet", 5, 15, "Raf B-2", "HSS", "Bosch"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="urun_sablonu.xlsx"'},
    )


@api.post("/products/import")
async def product_import(file: UploadFile = File(...), commit: bool = False,
                         user=Depends(require_admin)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyası yükleyin")
    content = await file.read()
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Dosyada veri yok (başlık + en az bir satır gerekli)")

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h}
    for r in ("name", "category"):
        if r not in idx:
            raise HTTPException(status_code=400, detail=f"Zorunlu sütun eksik: {r}")

    def _get(row, col, default=None):
        i = idx.get(col)
        if i is None or i >= len(row):
            return default
        return row[i]

    def _safe_str(v):
        if v is None:
            return ""
        return str(v).strip().lstrip("\ufeff")

    def _safe_num(v, default=0.0):
        if v is None or (isinstance(v, str) and not v.strip()):
            return default
        try:
            if isinstance(v, str):
                v = v.replace(",", ".").strip()
            x = float(v)
        except (TypeError, ValueError):
            raise
        if x != x or x == float("inf") or x == float("-inf"):
            return default
        return x

    preview = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or _safe_str(v) == "" for v in row):
            continue
        code = _safe_str(_get(row, "code"))
        name = _safe_str(_get(row, "name"))
        category = _safe_str(_get(row, "category"))
        if not name or not category:
            preview.append({"row": row_num, "action": "skip",
                            "error": "name/category boş olamaz",
                            "data": {"code": code, "name": name, "category": category}})
            continue
        try:
            unit = _safe_str(_get(row, "unit")) or "adet"
            min_stock = _safe_num(_get(row, "min_stock"), 0.0)
            current_stock = _safe_num(_get(row, "current_stock"), 0.0)
        except (TypeError, ValueError):
            preview.append({"row": row_num, "action": "skip",
                            "error": "min_stock / current_stock sayı olmalı",
                            "data": {"code": code}})
            continue
        location = _safe_str(_get(row, "location"))
        quality = _safe_str(_get(row, "quality"))
        brand = _safe_str(_get(row, "brand"))
        image_url = _safe_str(_get(row, "image_url"))
        existing = await db.products.find_one({"code": code}) if code else None
        action = "update" if existing else "create"
        preview.append({
            "row": row_num, "action": action, "error": None,
            "data": {"code": code, "name": name, "category": category, "unit": unit,
                     "min_stock": min_stock, "current_stock": current_stock,
                     "location": location, "quality": quality, "brand": brand, "image_url": image_url},
        })

    stats = {
        "total": len(preview),
        "create": sum(1 for p in preview if p["action"] == "create"),
        "update": sum(1 for p in preview if p["action"] == "update"),
        "skip": sum(1 for p in preview if p["action"] == "skip"),
    }

    if not commit:
        return {"committed": False, "stats": stats, "preview": preview}

    created, updated = 0, 0
    for p in preview:
        if p["action"] == "skip":
            continue
        d = p["data"]
        if p["action"] == "create":
            if not d.get("code"):
                d["code"] = await _next_product_code()
            await db.products.insert_one({**d, "id": new_id(),
                                          "created_at": now_utc().isoformat()})
            created += 1
        else:
            await db.products.update_one({"code": d["code"]}, {"$set": d})
            updated += 1
    return {"committed": True, "stats": stats, "created": created,
            "updated": updated, "preview": preview}


# ==================== Daily digest email + scheduler ====================
TR_TZ = ZoneInfo("Europe/Istanbul")


async def send_daily_digest():
    """Send yesterday's stock in/out summary to OWNER_EMAIL."""
    if not EMAIL_KEY or not OWNER_EMAIL:
        logger.info("Daily digest: email not configured, skipping")
        return

    now_tr = datetime.now(TR_TZ)
    yesterday = (now_tr - timedelta(days=1)).date()
    day_start = datetime.combine(yesterday, datetime.min.time(), tzinfo=TR_TZ).astimezone(timezone.utc).isoformat()
    day_end = datetime.combine(yesterday, datetime.max.time(), tzinfo=TR_TZ).astimezone(timezone.utc).isoformat()

    movements = await db.movements.find(
        {"created_at": {"$gte": day_start, "$lte": day_end}}, {"_id": 0}
    ).to_list(5000)

    ins = [m for m in movements if m.get("type") == "in"]
    outs = [m for m in movements if m.get("type") == "out"]
    in_total = round(sum(m.get("total", 0) for m in ins), 2)
    out_total = round(sum(m.get("total", 0) for m in outs), 2)

    def _rows(items):
        if not items:
            return '<tr><td colspan="4" style="padding:12px; color:#94a3b8; text-align:center;">Kayıt yok</td></tr>'
        return "".join(
            f'<tr><td style="padding:8px 12px; border-top:1px solid #334155;">{m.get("product_code","")}</td>'
            f'<td style="padding:8px 12px; border-top:1px solid #334155;">{m.get("product_name","")}</td>'
            f'<td style="padding:8px 12px; border-top:1px solid #334155; text-align:right; color:#e2e8f0;">{m.get("quantity",0)}</td>'
            f'<td style="padding:8px 12px; border-top:1px solid #334155; text-align:right; color:#e2e8f0;">₺{m.get("total",0):,.2f}</td></tr>'
            for m in items
        )

    subject = f"[Günlük Özet] {yesterday.strftime('%d.%m.%Y')} — Takımhane Hareketleri"
    html = f"""
    <div style="font-family: Arial, sans-serif; max-width:720px; margin:0 auto; background:#0f172a; color:#f8fafc; padding:24px; border-radius:12px;">
      <h2 style="color:#60a5fa; margin:0 0 4px 0; font-size:22px;">Günlük Takımhane Özeti</h2>
      <div style="color:#94a3b8; font-size:13px; margin-bottom:20px;">{yesterday.strftime('%d %B %Y')}</div>

      <div style="display:flex; gap:12px; margin-bottom:24px;">
        <div style="flex:1; background:#1e293b; padding:16px; border-radius:8px; border-left:3px solid #10b981;">
          <div style="color:#94a3b8; font-size:11px; text-transform:uppercase; letter-spacing:1px;">Giriş</div>
          <div style="font-size:24px; font-weight:bold; color:#10b981;">{len(ins)}</div>
          <div style="color:#cbd5e1; font-size:13px;">₺{in_total:,.2f}</div>
        </div>
        <div style="flex:1; background:#1e293b; padding:16px; border-radius:8px; border-left:3px solid #f59e0b;">
          <div style="color:#94a3b8; font-size:11px; text-transform:uppercase; letter-spacing:1px;">Çıkış</div>
          <div style="font-size:24px; font-weight:bold; color:#f59e0b;">{len(outs)}</div>
          <div style="color:#cbd5e1; font-size:13px;">₺{out_total:,.2f}</div>
        </div>
      </div>

      <h3 style="color:#10b981; font-size:15px; margin:16px 0 8px 0;">Stok Girişleri</h3>
      <table style="width:100%; border-collapse:collapse; background:#1e293b; border-radius:6px; overflow:hidden; font-size:13px;">
        <thead><tr style="background:#0f172a;">
          <th style="padding:10px 12px; text-align:left; color:#94a3b8;">Kod</th>
          <th style="padding:10px 12px; text-align:left; color:#94a3b8;">Ürün</th>
          <th style="padding:10px 12px; text-align:right; color:#94a3b8;">Miktar</th>
          <th style="padding:10px 12px; text-align:right; color:#94a3b8;">Tutar</th>
        </tr></thead>
        <tbody>{_rows(ins)}</tbody>
      </table>

      <h3 style="color:#f59e0b; font-size:15px; margin:20px 0 8px 0;">Stok Çıkışları</h3>
      <table style="width:100%; border-collapse:collapse; background:#1e293b; border-radius:6px; overflow:hidden; font-size:13px;">
        <thead><tr style="background:#0f172a;">
          <th style="padding:10px 12px; text-align:left; color:#94a3b8;">Kod</th>
          <th style="padding:10px 12px; text-align:left; color:#94a3b8;">Ürün</th>
          <th style="padding:10px 12px; text-align:right; color:#94a3b8;">Miktar</th>
          <th style="padding:10px 12px; text-align:right; color:#94a3b8;">Tutar</th>
        </tr></thead>
        <tbody>{_rows(outs)}</tbody>
      </table>

      <p style="color:#94a3b8; font-size:12px; margin-top:24px;">CNC Takımhane Stok Takip Sistemi — Otomatik günlük özet</p>
    </div>
    """
    payload = {"to": [OWNER_EMAIL], "subject": subject, "html": html, "from_name": EMAIL_FROM_NAME}
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.post(f"{EMAIL_BASE_URL}/api/v1/email/send",
                             headers={"X-Email-Key": EMAIL_KEY}, json=payload)
            r.raise_for_status()
        logger.info(f"Daily digest sent for {yesterday}")
    except Exception as e:
        logger.error(f"Daily digest email failed: {e}")


@api.post("/admin/send-daily-digest")
async def trigger_daily_digest(user=Depends(require_admin)):
    """Manual trigger for daily digest — useful for testing."""
    await send_daily_digest()
    return {"ok": True}


_scheduler: Optional[AsyncIOScheduler] = None


@app.on_event("startup")
async def start_scheduler():
    global _scheduler
    if _scheduler is None:
        _scheduler = AsyncIOScheduler(timezone=TR_TZ)
        _scheduler.add_job(send_daily_digest, CronTrigger(hour=8, minute=0, timezone=TR_TZ),
                           id="daily_digest", replace_existing=True)
        _scheduler.start()
        logger.info("Scheduler started: daily digest at 08:00 Europe/Istanbul")


@app.on_event("shutdown")
async def stop_scheduler():
    global _scheduler
    if _scheduler:
        _scheduler.shutdown(wait=False)
        _scheduler = None


app.include_router(api)
